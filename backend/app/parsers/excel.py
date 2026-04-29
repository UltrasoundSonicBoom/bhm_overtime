"""Excel/CSV 결정론적 파서 — openpyxl + pandas."""
import io
import re
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from app.parsers.duty_codes import map_duty_code, map_duty_confidence
from app.schemas.schedule import DutyGrid, ScheduleRow


def parse_excel_bytes(data: bytes, file_name: Optional[str] = None) -> DutyGrid:
    """Excel 바이트 → DutyGrid."""
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as e:
        return _empty_grid(f"openpyxl_error: {e}", "excel-v1.0", "excel")

    sheet_names = wb.sheetnames
    if not sheet_names:
        return _empty_grid("no_sheets", "excel-v1.0", "excel")

    sheet_name = sheet_names[0]
    ws = wb[sheet_name]
    grid = []
    for row in ws.iter_rows(values_only=True):
        grid.append([("" if c is None else str(c)) for c in row])

    return _parse_grid(grid, sheet_name, file_name, parser_version="excel-v1.0", source="excel")


def parse_csv_text(text: str, file_name: Optional[str] = None) -> DutyGrid:
    """CSV 텍스트 → DutyGrid."""
    try:
        df = pd.read_csv(io.StringIO(text), header=None, dtype=str, keep_default_na=False)
    except Exception as e:
        return _empty_grid(f"pandas_error: {e}", "csv-v1.0", "csv")

    grid = df.fillna("").astype(str).values.tolist()
    return _parse_grid(grid, file_name or "", file_name, parser_version="csv-v1.0", source="csv")


def _parse_grid(
    grid: list[list[str]],
    title_hint: str,
    file_name: Optional[str],
    parser_version: str,
    source: str,
) -> DutyGrid:
    if not grid:
        return _empty_grid("empty_sheet", parser_version, source)

    header = _find_day_header_row(grid)
    if header is None:
        return _empty_grid("no_day_header", parser_version, source)

    header_row, day_map = header
    name_col = _find_name_column(grid, header_row)

    rows: list[ScheduleRow] = []
    total_conf_sum = 0.0
    total_cell_count = 0
    unmapped_count = 0

    for r in range(header_row + 1, len(grid)):
        row = grid[r]
        if not row:
            continue
        name = (row[name_col] if name_col < len(row) else "").strip()
        if not name or _looks_like_metadata(name):
            continue

        days: dict[str, str] = {}
        for col, day_num in day_map.items():
            raw = row[col] if col < len(row) else ""
            code = map_duty_code(raw)
            conf = map_duty_confidence(raw)
            if raw and str(raw).strip():
                total_conf_sum += conf
                total_cell_count += 1
                if conf == 0.0:
                    unmapped_count += 1
            days[str(day_num)] = code

        if days:
            rows.append(ScheduleRow(name=name, days=days))

    confidence = 1.0 if total_cell_count == 0 else max(0.0, total_conf_sum / total_cell_count)
    notes = f"매핑 실패 {unmapped_count}셀" if unmapped_count > 0 else ""
    month = _infer_month(title_hint) or _infer_month(file_name) or None
    dept = _infer_dept(title_hint) or _infer_dept(file_name) or None

    return DutyGrid(
        month=month,
        dept=dept,
        rows=rows,
        confidence=confidence,
        notes=notes,
        parser_version=parser_version,
        source=source,
    )


def _find_day_header_row(grid: list[list[str]]) -> Optional[tuple[int, dict[int, int]]]:
    best_row = -1
    best_day_map: dict[int, int] = {}
    best_score = 0

    max_rows = min(5, len(grid))
    for r in range(max_rows):
        row = grid[r]
        if not row:
            continue
        day_map: dict[int, int] = {}
        score = 0
        for c, cell in enumerate(row):
            cell_str = str(cell or "").strip()
            m = re.match(r"^(\d{1,2})(일)?$", cell_str)
            if m:
                day = int(m.group(1))
                if 1 <= day <= 31:
                    day_map[c] = day
                    score += 1
        if score > best_score:
            best_score = score
            best_row = r
            best_day_map = day_map

    if best_score < 7:
        return None
    return (best_row, best_day_map)


def _find_name_column(grid: list[list[str]], header_row: int) -> int:
    korean_name_re = re.compile(r"^[가-힣]{2,5}$")
    col_counts: dict[int, int] = {}
    end = min(header_row + 20, len(grid))
    for r in range(header_row + 1, end):
        row = grid[r]
        if not row:
            continue
        for c in range(min(len(row), 5)):
            cell = str(row[c] or "").strip()
            if korean_name_re.match(cell):
                col_counts[c] = col_counts.get(c, 0) + 1

    if not col_counts:
        return 0
    return max(col_counts, key=lambda k: col_counts[k])


META_TOKENS = ("합계", "비고", "총", "소계", "Total", "Sum", "합", "총계")


def _looks_like_metadata(name: str) -> bool:
    return any(tok in name for tok in META_TOKENS)


def _infer_month(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    m1 = re.search(r"(\d{4})[\-.년\s]+(\d{1,2})월?", text)
    if m1:
        return f"{m1.group(1)}-{int(m1.group(2)):02d}"
    m2 = re.search(r"(\d{2})[\-.](\d{1,2})", text)
    if m2:
        return f"20{m2.group(1)}-{int(m2.group(2)):02d}"
    return None


DEPT_WHITELIST = ("ICU", "CCU", "NICU", "응급실", "병동", "수술실", "외래")


def _infer_dept(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    for d in DEPT_WHITELIST:
        if d in text:
            return d
    return None


def _empty_grid(reason: str, parser_version: str, source: str) -> DutyGrid:
    return DutyGrid(
        month=None,
        dept=None,
        rows=[],
        confidence=0.0,
        notes=reason,
        parser_version=parser_version,
        source=source,
    )
