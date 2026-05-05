from __future__ import annotations

import argparse
import csv
import datetime as dt
import html
import json
import re
import shutil
import tempfile
import time
from pathlib import Path
from typing import Any, Iterable


PROJECT_DIR = Path(__file__).resolve().parent
DEFAULT_OUTPUT = PROJECT_DIR / "output"

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
CSV_EXTENSIONS = {".csv", ".tsv"}
XLSX_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}

EXCEL_PAYMENT_LABEL_ROWS = range(4, 10)
EXCEL_DEDUCTION_LABEL_ROWS = range(17, 23)
EXCEL_AMOUNT_OFFSET = 6
EXCEL_LABEL_COLS = [2, 6, 10, 14, 15, 17, 18, 20, 23, 26, 27, 29, 30]
EXCEL_DEDUCTION_AMOUNT_COLS = [2, 6, 10, 14, 15, 17, 18]
EXCEL_WORK_METRIC_COLS = [20, 23, 26, 27, 29, 30]

PAY_GROUPS = [
    ("기본급여", ["기본기준급", "근속가산기본급", "능력급", "상여금", "특별상여금", "조정급", "승급호봉분"]),
    ("수당", ["가계지원비", "정근수당", "명절지원비", "직책수당", "장기근속수당", "가족수당", "군복무수당", "경력인정수당", "업무보조비"]),
    ("보조비", ["급식보조비", "교통보조비", "의학연구지원금"]),
    ("진료·연구", ["의업수당", "진료수당", "임상연구비", "연구실습비", "연구보조비", "의학연구비", "진료기여수당", "선택진료수당", "보직교수기여수당", "진료비보조", "연구장려수당", "진료지원수당"]),
    ("성과·기타", ["성과급", "기타수당", "별정수당(직무)", "별정수당(약제부+전문간호사+기타)", "산전후보전급여", "당직비", "주치의수당", "원외근무수당", "육아휴직수당", "간호간병특별수당"]),
    ("초과근무", ["시간외수당", "휴일수당", "야간수당", "통상야간", "명절수당", "법정공휴일수당", "야간근무가산금", "대체근무가산금", "대체근무통상야근수당"]),
    ("연차·휴가", ["연차수당", "연차보전수당", "무급생휴공제", "무급가족돌봄휴가", "무급난임휴가", "육아기근로시간단축", "기타지급1", "기타지급2"]),
]

DEDUCTION_GROUPS = [
    ("세금", ["소득세", "주민세", "농특세", "소득세(정산)", "주민세(정산)", "농특세(정산)"]),
    ("4대보험", ["국민건강", "장기요양", "국민연금", "고용보험", "국민건강(정산)", "장기요양(정산)", "국민연금(정산)", "고용보험(정산)"]),
    ("연금·대출", ["교원장기급여", "교원대출상환", "사학연금부담금", "사학연금대여상환금", "사학연금정산금", "마을금고상환", "대학학자금대출상환", "채권가압류"]),
    ("조합·기금", ["노동조합비", "노조기금", "병원발전기금", "후원회비", "상조회비", "의사협회비", "기금협의회비", "전공의협회비", "전공의동창회비", "기금출연금", "장학지원금공제", "의국비"]),
    ("생활비", ["주차료", "기숙사비", "보육료", "식대공제", "기타공제1", "기타공제2", "기타공제3"]),
]

PDF_WIDTH = 841.0
PDF_HEIGHT = 595.0
PDF_X_BOUNDS = [72, 143, 201, 263, 325, 387, 448, 510, 572, 634, 696, 758]
PDF_PAYMENT_Y_CENTERS = [146, 159.5, 173.1, 186.7, 200.3, 213.9]
PDF_DEDUCTION_Y_CENTERS = [311.4, 325.0, 338.6, 352.2, 365.8, 379.4]

HEADER_BBOXES = {
    "title": [260, 0, 510, 30],
    "pay_date": [670, 5, 752, 25],
    "personal_number": [88, 32, 123, 52],
    "name": [209, 32, 250, 52],
    "job_category": [335, 32, 370, 52],
    "pay_step": [456, 32, 494, 52],
    "department": [578, 32, 620, 52],
    "hire_date": [698, 32, 748, 52],
}

SUPPLEMENTAL_IMAGE_BBOXES = {
    "teacher_mutual_aid_member_number": [105, 392, 155, 414],
    "payroll_account_change_instruction": [25, 434, 570, 454],
    "contact_general": [680, 398, 750, 418],
    "contact_short_time": [680, 426, 750, 446],
    "contact_doctor_environment": [675, 448, 750, 484],
    "contact_hr_topic": [580, 484, 675, 505],
    "contact_hr": [680, 484, 750, 505],
}

PAYMENT_LABEL_GRID = [
    [
        "기본기준급",
        "정근수당",
        "연구보조비",
        "진료기여수당(협진)",
        "성과급",
        "급식보조비",
        "시간외수당",
        "야간근무가산금",
        "당직비",
        "기타지급1",
        "대체근무가산금",
    ],
    [
        "근속가산기본급",
        "명절지원비",
        "의학연구비",
        "진료비보조",
        "기타수당",
        "교통보조비",
        "휴일수당",
        "무급생휴공제",
        "주치의수당",
        "기타지급2",
        "대체근무 통상야근수당",
    ],
    [
        "능력급",
        "의업수당",
        "진료기여수당",
        "",
        "조정급",
        "직책수당",
        "야간수당",
        "별정수당(약제부+전문간호사+기타)",
        "군복무수당",
        "간호간병특별수당",
        "",
    ],
    [
        "상여금",
        "진료수당",
        "선택진료수당",
        "별정수당(직무)",
        "승급호봉분",
        "업무보조비",
        "통상야간",
        "산전후보전급여",
        "육아휴직수당",
        "연차수당",
        "급여총액",
    ],
    [
        "특별상여금",
        "임상연구비",
        "보직교수기여수당",
        "연구장려수당",
        "경력인정수당",
        "장기근속수당",
        "명절수당",
        "가족수당",
        "무급가족돌봄휴가",
        "연차보전수당",
        "공제총액",
    ],
    [
        "가계지원비",
        "연구실습비",
        "진료기여수당(토요진료)",
        "진료지원수당",
        "의학연구지원금",
        "원외근무수당",
        "법정공휴일수당",
        "",
        "육아기근로시간단축",
        "무급난임휴가",
        "실지급액",
    ],
]

DEDUCTION_LABEL_GRID = [
    [
        "소득세",
        "국민건강",
        "고용보험",
        "장학지원금공제",
        "병원발전기금",
        "전공의협회비",
        "식대공제",
        "총근로시간",
        "시간외근무시간",
        "야간근무가산횟수",
        "무급생휴일",
    ],
    [
        "주민세",
        "장기요양",
        "고용보험(정산)",
        "노동조합비",
        "후원회비",
        "전공의동창회비",
        "대학학자금대출상환",
        "통상근로시간",
        "휴일근무시간",
        "대체근무가산횟수",
        "지급연차갯수",
    ],
    [
        "농특세",
        "국민연금",
        "교원장기급여",
        "노조기금",
        "의국비",
        "기금출연금",
        "기타공제1",
        "야간근로시간",
        "야간근무시간",
        "대체근무통상야근시간",
        "사용연차",
    ],
    [
        "소득세(정산)",
        "국민건강(정산)",
        "교원대출상환",
        "주차료",
        "상조회비",
        "사학연금부담금",
        "기타공제2",
        "주휴시간",
        "통상야근시간",
        "",
        "발생연차",
    ],
    [
        "주민세(정산)",
        "장기요양(정산)",
        "마을금고상환",
        "기숙사비",
        "의사협회비",
        "사학연금대여상환금",
        "기타공제3",
        "유급휴일",
        "명절근무시간",
        "",
        "",
    ],
    [
        "농특세(정산)",
        "국민연금(정산)",
        "채권가압류",
        "보육료",
        "기금협의회비",
        "사학연금정산금",
        "",
        "",
        "법정공휴일근무시간",
        "",
        "",
    ],
]


def now_iso() -> str:
    return dt.datetime.now().isoformat(timespec="seconds")


def detect_payroll_file_kind(path: Path) -> str:
    suffix = path.suffix.lower()
    prefix = path.read_bytes()[:4096]
    stripped = prefix.lstrip().lower()
    if prefix.startswith(b"\xff\xd8\xff") or prefix.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image"
    if prefix.startswith(b"%PDF"):
        return "pdf"
    if prefix.startswith(b"\xd0\xcf\x11\xe0"):
        return "xls-binary"
    if prefix.startswith(b"PK\x03\x04"):
        return "xlsx"
    if suffix in IMAGE_EXTENSIONS:
        return "image"
    if suffix in CSV_EXTENSIONS or b"section,label,amount" in stripped:
        return "csv"
    if suffix in XLSX_EXTENSIONS:
        return "xlsx"
    return "unknown"


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n").strip()
    text = re.sub(r"\s+", " ", text)
    if text.endswith(".0") and text[:-2].lstrip("-").isdigit():
        return text[:-2]
    return text


def compact_text(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def parse_money(value: Any) -> int | None:
    text = clean_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace("원", "").replace("￦", "").strip()
    if re.fullmatch(r"-?\d+(?:\.\d+)?", text):
        return int(round(float(text)))
    match = re.search(r"-?\d[\d,]*(?:\.\d+)?", clean_text(value))
    if match:
        return int(round(float(match.group(0).replace(",", ""))))
    return None


def excel_date(value: Any, datemode: int) -> str:
    if isinstance(value, float | int):
        try:
            import xlrd

            return xlrd.xldate.xldate_as_datetime(value, datemode).date().isoformat()
        except Exception:
            return clean_text(value)
    return clean_text(value)


def excel_serial_date(value: Any) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    amount = parse_money(value)
    if amount is not None and 30000 <= amount <= 60000:
        return (dt.date(1899, 12, 30) + dt.timedelta(days=amount)).isoformat()
    return clean_text(value)


def pay_period_from_title(title: str) -> str | None:
    compact = compact_text(title)
    match = re.search(r"(\d{4})년도?(\d{1,2})월분", compact)
    if not match:
        return None
    return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}"


def statement_kind_from_title(title: str) -> str:
    match = re.search(r"\(([^)]+)\)", title)
    return match.group(1) if match else "regular"


def build_item(label: str, amount: int | None, raw_value: Any, source: str, row: int | None = None, col: int | None = None) -> dict[str, Any]:
    return {
        "label": clean_text(label),
        "amount": amount,
        "raw_value": clean_text(raw_value),
        "source": source,
        "row": row,
        "col": col,
    }


def parse_xls(path: Path) -> dict[str, Any]:
    import xlrd

    book = xlrd.open_workbook(str(path), formatting_info=False)
    sheet = book.sheet_by_index(0)
    title = first_cell_matching(sheet, "급여명세서") or path.stem
    pay_date = excel_date(value_right_of_label(sheet, "급여지급일"), book.datemode)
    employee = {
        "personal_number": clean_text(value_right_of_label(sheet, "개인번호")),
        "name": clean_text(value_right_of_label(sheet, "성 명") or value_right_of_label(sheet, "성명")),
        "job_category": clean_text(value_right_of_label(sheet, "직 종") or value_right_of_label(sheet, "직종")),
        "pay_step": clean_text(value_right_of_label(sheet, "급여연차")),
        "department": clean_text(value_right_of_label(sheet, "소 속") or value_right_of_label(sheet, "소속")),
        "hire_date": clean_text(value_right_of_label(sheet, "입사년월")),
    }
    pay_anchor = find_anchor_row(sheet, ["기본기준급"])
    deduction_anchor = find_anchor_row(sheet, ["소득세", "국민건강"])
    earnings = xls_items_from_anchor(sheet, pay_anchor, EXCEL_LABEL_COLS, "earning") if pay_anchor is not None else xls_items(sheet, EXCEL_PAYMENT_LABEL_ROWS, "earning", EXCEL_LABEL_COLS)
    deductions = (
        xls_items_from_anchor(sheet, deduction_anchor, EXCEL_DEDUCTION_AMOUNT_COLS, "deduction")
        if deduction_anchor is not None
        else xls_items(sheet, EXCEL_DEDUCTION_LABEL_ROWS, "deduction", EXCEL_DEDUCTION_AMOUNT_COLS)
    )
    work_metrics = (
        xls_items_from_anchor(sheet, deduction_anchor, EXCEL_WORK_METRIC_COLS, "work_metric")
        if deduction_anchor is not None
        else xls_items(sheet, EXCEL_DEDUCTION_LABEL_ROWS, "work_metric", EXCEL_WORK_METRIC_COLS)
    )
    supplemental = merge_supplemental(
        parse_page1_supplement_from_sheet(sheet),
        parse_page2_sheet(book),
        {"work_metrics": work_metrics},
    )
    return build_statement(path, "xls-binary", title, pay_date, employee, earnings, deductions, supplemental)


def first_cell_matching(sheet: Any, needle: str) -> str | None:
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            value = clean_text(sheet.cell_value(row, col))
            if needle in value:
                return value
    return None


def value_right_of_label(sheet: Any, label: str) -> Any:
    label_compact = compact_text(label).rstrip(":：")
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell_compact = compact_text(sheet.cell_value(row, col)).rstrip(":：")
            if cell_compact == label_compact:
                for next_col in range(col + 1, sheet.ncols):
                    value = sheet.cell_value(row, next_col)
                    if clean_text(value):
                        return value
    return ""


def value_right_of_label_limited(sheet: Any, label: str, max_gap: int = 4) -> Any:
    label_compact = compact_text(label).rstrip(":：")
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            cell_compact = compact_text(sheet.cell_value(row, col)).rstrip(":：")
            if cell_compact == label_compact:
                for next_col in range(col + 1, min(sheet.ncols, col + max_gap + 1)):
                    value = sheet.cell_value(row, next_col)
                    if clean_text(value):
                        return value
    return ""


def find_anchor_row(sheet: Any, needles: list[str]) -> int | None:
    needle_compacts = [compact_text(needle) for needle in needles]
    for row in range(sheet.nrows):
        row_text = "|".join(compact_text(sheet.cell_value(row, col)) for col in range(sheet.ncols))
        if all(needle in row_text for needle in needle_compacts):
            return row
    return None


def xls_items(sheet: Any, label_rows: Iterable[int], source: str, columns: Iterable[int] | None = None) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    selected_columns = list(columns) if columns is not None else list(range(sheet.ncols))
    for label_row in label_rows:
        amount_row = label_row + EXCEL_AMOUNT_OFFSET
        if amount_row >= sheet.nrows:
            continue
        for col in selected_columns:
            if col >= sheet.ncols:
                continue
            label = clean_text(sheet.cell_value(label_row, col))
            if not is_item_label(label):
                continue
            raw_value = sheet.cell_value(amount_row, col)
            amount = parse_money(raw_value)
            if amount is None and not is_total_label(label):
                continue
            items.append(build_item(label, amount, raw_value, source, amount_row, col))
    return items


def xls_items_from_anchor(sheet: Any, anchor_row: int, columns: Iterable[int], source: str) -> list[dict[str, Any]]:
    return xls_items(sheet, range(anchor_row, anchor_row + 6), source, columns)


def is_item_label(label: str) -> bool:
    if not label:
        return False
    compact = compact_text(label)
    if compact in {"지급내역", "공제내역"}:
        return False
    return bool(re.search(r"[가-힣A-Za-z]", compact))


def is_total_label(label: str) -> bool:
    return compact_text(label) in {"급여총액", "공제총액", "실지급액"}


def parse_page2(book: Any) -> dict[str, Any]:
    if book.nsheets < 2:
        return {}
    sheet = book.sheet_by_index(1)
    values: dict[str, Any] = {}
    for row in range(sheet.nrows):
        cells = [clean_text(sheet.cell_value(row, col)) for col in range(sheet.ncols)]
        for idx, value in enumerate(cells):
            if not value or idx + 1 >= len(cells):
                continue
            if value in {"개인번호", "이름"}:
                continue
            next_value = cells[idx + 1]
            if next_value:
                values[value] = parse_money(next_value) if re.search(r"\d", next_value) else next_value
    return values


def empty_supplemental() -> dict[str, Any]:
    return {
        "membership": {
            "teacher_mutual_aid_member_number": "",
            "private_school_pension_institution_number": "",
        },
        "work_basis": {},
        "work_metrics": [],
        "calculation_details": [],
        "messages": {
            "family_allowance_target": "",
            "payroll_account_change_instruction": "",
        },
        "contacts": [],
    }


def merge_supplemental(*parts: dict[str, Any]) -> dict[str, Any]:
    merged = empty_supplemental()
    for part in parts:
        if not part:
            continue
        for section in ["membership", "work_basis", "messages"]:
            for key, value in part.get(section, {}).items():
                if value not in (None, ""):
                    merged[section][key] = value
        merged["work_metrics"].extend(part.get("work_metrics", []))
        merged["calculation_details"].extend(part.get("calculation_details", []))
        merged["contacts"].extend(part.get("contacts", []))
    merged["work_metrics"] = dedupe_dicts(merged["work_metrics"], ["label", "amount", "row", "col"])
    merged["calculation_details"] = dedupe_dicts(merged["calculation_details"], ["section", "item", "formula", "amount"])
    merged["contacts"] = dedupe_dicts(merged["contacts"], ["topic", "team", "category", "extension"])
    return merged


def dedupe_dicts(items: list[dict[str, Any]], keys: list[str]) -> list[dict[str, Any]]:
    seen: set[tuple[Any, ...]] = set()
    result: list[dict[str, Any]] = []
    for item in items:
        marker = tuple(item.get(key) for key in keys)
        if marker in seen:
            continue
        seen.add(marker)
        result.append(item)
    return result


def parse_page1_supplement_from_sheet(sheet: Any) -> dict[str, Any]:
    supplemental = supplemental_from_text_lines(sheet_text_lines(sheet))
    teacher_number = clean_text(value_right_of_label_limited(sheet, "교원공제회원번호"))
    pension_number = clean_text(value_right_of_label_limited(sheet, "사학연금기관번호"))
    if teacher_number:
        supplemental["membership"]["teacher_mutual_aid_member_number"] = teacher_number
    if pension_number and "급여문의" not in pension_number:
        supplemental["membership"]["private_school_pension_institution_number"] = pension_number
    return supplemental


def sheet_text_lines(sheet: Any) -> list[str]:
    lines: list[str] = []
    for row in range(sheet.nrows):
        for col in range(sheet.ncols):
            value = sheet.cell_value(row, col)
            if clean_text(value):
                lines.extend(normalize_text_lines([value]))
    return lines


def normalize_text_lines(values: Iterable[Any]) -> list[str]:
    lines: list[str] = []
    for value in values:
        text = "" if value is None else str(value).replace("<br>", "\n").replace("\r", "\n")
        for part in text.split("\n"):
            cleaned = clean_text(part)
            if cleaned:
                lines.append(cleaned)
    return lines


def supplemental_from_text_lines(raw_lines: Iterable[Any]) -> dict[str, Any]:
    lines = normalize_text_lines(raw_lines)
    supplemental = empty_supplemental()
    supplemental["membership"]["teacher_mutual_aid_member_number"] = value_after_label_in_lines(lines, "교원공제회원번호")
    pension_number = value_after_label_in_lines(lines, "사학연금기관번호")
    if pension_number and "급여문의" not in pension_number:
        supplemental["membership"]["private_school_pension_institution_number"] = pension_number
    supplemental["messages"]["family_allowance_target"] = value_after_label_in_lines(lines, "가족수당지급대상", allow_next=False)
    supplemental["messages"]["payroll_account_change_instruction"] = value_after_label_in_lines(lines, "급여계좌변경")
    supplemental["contacts"] = parse_contact_lines(lines)
    return supplemental


def value_after_label_in_lines(lines: list[str], label: str, allow_next: bool = True) -> str:
    label_compact = compact_text(label).rstrip(":：")
    stop_labels = {"개인번호", "성명", "직종", "급여연차", "소속", "입사년월", "사학연금기관번호", "급여문의", "귀하의노고에진심으로감사드립니다"}
    for index, line in enumerate(lines):
        compact = compact_text(line).rstrip(":：")
        if not compact.startswith(label_compact):
            continue
        rest = clean_text(re.sub(rf"^{re.escape(label)}\s*[:：]?\s*", "", line))
        if rest and compact_text(rest).rstrip(":：") != label_compact:
            return rest
        if allow_next and index + 1 < len(lines):
            next_line = lines[index + 1]
            if compact_text(next_line).rstrip(":：") not in stop_labels:
                return next_line
    return ""


def parse_contact_lines(lines: list[str]) -> list[dict[str, str]]:
    contacts: list[dict[str, str]] = []
    welfare_team = "복지팀" if any(compact_text(line) == "복지팀" for line in lines) else ""

    for line in lines:
        compact = compact_text(line)
        match = re.fullmatch(r"(일반직|단시간)(\d{3,4})", compact)
        if match:
            contacts.append({"topic": "급여문의", "team": welfare_team, "category": match.group(1), "extension": match.group(2)})

    for index, line in enumerate(lines):
        if compact_text(line) == "2731":
            labels = []
            for previous in lines[max(0, index - 2):index]:
                if not re.search(r"\d", previous):
                    labels.append(previous)
            if labels:
                category = clean_text(" ".join(labels)).replace("일사직/", "의사직/")
                contacts.append({"topic": "급여문의", "team": welfare_team, "category": category, "extension": "2731"})

        compact = compact_text(line)
        match = re.fullmatch(r"인사팀(\d{3,4})", compact)
        if match:
            topic = lines[index - 1] if index > 0 else "휴직,연차,연장근무"
            contacts.append({"topic": clean_text(topic), "team": "인사팀", "category": "", "extension": match.group(1)})

    return dedupe_dicts(contacts, ["topic", "team", "category", "extension"])


def supplemental_from_image_raw(raw: dict[str, str]) -> dict[str, Any]:
    supplemental = empty_supplemental()
    member_raw = raw.get("teacher_mutual_aid_member_number", "")
    member_match = re.search(r"[A-Z]\d{3,}", member_raw)
    if member_match:
        supplemental["membership"]["teacher_mutual_aid_member_number"] = member_match.group(0)
    supplemental["messages"]["payroll_account_change_instruction"] = clean_text(raw.get("payroll_account_change_instruction"))
    contact_lines = [
        "급여문의",
        "복지팀",
        raw.get("contact_general", ""),
        raw.get("contact_short_time", ""),
        raw.get("contact_doctor_environment", ""),
        raw.get("contact_hr_topic", ""),
        raw.get("contact_hr", ""),
    ]
    supplemental["contacts"] = parse_contact_lines(normalize_text_lines(contact_lines))
    return supplemental


def parse_page2_sheet(book: Any) -> dict[str, Any]:
    if book.nsheets < 2:
        return {}
    sheet = book.sheet_by_index(1)
    return {
        **empty_supplemental(),
        "work_basis": {
            "personal_number": only_digits(value_right_of_label(sheet, "개인번호")),
            "name": compact_text(value_right_of_label(sheet, "이름")),
            "work_days": parse_money(value_right_of_label(sheet, "근로일수")),
            "previous_month_regular_wage": parse_money(value_right_of_label(sheet, "전월통상임금")),
            "annual_average_wage": parse_money(value_right_of_label(sheet, "연차평균임금")),
        },
        "calculation_details": calculation_details_from_sheet(sheet),
    }


def calculation_details_from_sheet(sheet: Any) -> list[dict[str, Any]]:
    details: list[dict[str, Any]] = []
    header_row = None
    for row in range(sheet.nrows):
        cells = [compact_text(sheet.cell_value(row, col)) for col in range(sheet.ncols)]
        if "구분" in cells and "항목" in cells and "계산방법" in cells and "지급액" in cells:
            header_row = row
            break
    if header_row is None:
        return details

    for row in range(header_row + 1, sheet.nrows):
        section = clean_text(sheet.cell_value(row, 1))
        item = clean_text(sheet.cell_value(row, 2))
        formula = clean_text(sheet.cell_value(row, 4))
        amount = parse_money(sheet.cell_value(row, 9))
        if not section or not item:
            continue
        details.append(build_calculation_detail(section, item, formula, amount))
    return details


def parse_page2_text(text: str) -> dict[str, Any]:
    lines = normalize_text_lines(text.splitlines())
    supplemental = empty_supplemental()
    supplemental["work_basis"] = {
        "personal_number": only_digits(value_after_label_in_lines(lines, "개인번호")),
        "name": compact_text(value_after_label_in_lines(lines, "이름")),
        "work_days": parse_money(value_after_label_in_lines(lines, "근로일수")),
        "previous_month_regular_wage": parse_money(value_after_label_in_lines(lines, "전월통상임금")),
        "annual_average_wage": parse_money(value_after_label_in_lines(lines, "연차평균임금")),
    }
    supplemental["calculation_details"] = calculation_details_from_lines(lines)
    return supplemental


def calculation_details_from_lines(lines: list[str]) -> list[dict[str, Any]]:
    try:
        start = next(index for index, line in enumerate(lines) if compact_text(line) == "지급액") + 1
    except StopIteration:
        return []
    details: list[dict[str, Any]] = []
    index = start
    while index + 3 < len(lines):
        section, item, formula, amount_text = lines[index:index + 4]
        amount = parse_money(amount_text)
        if compact_text(section) in {"공제", "지급"} and item and amount is not None:
            details.append(build_calculation_detail(section, item, formula, amount))
            index += 4
        else:
            index += 1
    return details


def build_calculation_detail(section: str, item: str, formula: str, amount: int | None) -> dict[str, Any]:
    section_compact = compact_text(section)
    section_key = "deduction" if section_compact == "공제" else "earning" if section_compact == "지급" else section_compact
    return {
        "section": section_key,
        "section_label": clean_text(section),
        "item": clean_text(item),
        "formula": clean_text(formula),
        "amount": amount,
    }


def parse_pdf(path: Path) -> dict[str, Any]:
    import pymupdf

    document = pymupdf.open(path)
    try:
        if len(document) == 0:
            raise ValueError("PDF has no pages")
        page = document[0]
        text = page.get_text()
        words = page.get_text("words")
        if len(words) < 20:
            return parse_pdf_page_image(path, page)
        title = title_from_text(text) or path.stem
        pay_date = value_after_line(text, "급여지급일") or ""
        employee = employee_from_text(text)
        earnings = template_items_from_words(words, PAYMENT_LABEL_GRID, PDF_PAYMENT_Y_CENTERS, "earning")
        deductions = template_items_from_words(words, DEDUCTION_LABEL_GRID, PDF_DEDUCTION_Y_CENTERS, "deduction")
        supplemental = supplemental_from_text_lines(text.splitlines())
        if len(document) > 1:
            supplemental = merge_supplemental(supplemental, parse_page2_text(document[1].get_text()))
    finally:
        document.close()
    return build_statement(path, "pdf-text", title, pay_date, employee, earnings, deductions, supplemental)


def parse_pdf_page_image(path: Path, page: Any) -> dict[str, Any]:
    import pymupdf

    matrix = pymupdf.Matrix(2, 2)
    pixmap = page.get_pixmap(matrix=matrix, alpha=False)
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(prefix="payroll_pdf_page_", suffix=".png", delete=False) as handle:
            temp_path = Path(handle.name)
        pixmap.save(str(temp_path))
        statement = parse_image(temp_path)
        statement["source_file"] = str(path)
        statement["source_format"] = "pdf-image-ocr"
        statement["metadata"]["rendered_from_pdf"] = True
        return statement
    finally:
        if temp_path and temp_path.exists():
            temp_path.unlink(missing_ok=True)


def title_from_text(text: str) -> str | None:
    for line in text.splitlines():
        if "급여명세서" in line:
            compact = compact_text(line)
            match = re.search(r"(\d{4})년도(\d{1,2})월분급여명세서(?:\(([^)]+)\))?", compact)
            if match:
                suffix = f"({match.group(3)})" if match.group(3) else ""
                return f"{match.group(1)}년도 {int(match.group(2)):02d}월분 급여명세서{suffix}"
            return clean_text(line)
    return None


def value_after_line(text: str, label: str) -> str | None:
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    label_compact = compact_text(label).rstrip(":")
    for idx, line in enumerate(lines):
        compact = compact_text(line).rstrip(":")
        if compact.startswith(label_compact):
            rest = clean_text(re.sub(re.escape(label), "", line).strip(": "))
            if rest:
                return rest
            if idx + 1 < len(lines):
                return lines[idx + 1]
    return None


def employee_from_text(text: str) -> dict[str, str]:
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]

    def after(label: str) -> str:
        label_key = compact_text(label)
        for idx, line in enumerate(lines):
            if compact_text(line) == label_key and idx + 1 < len(lines):
                return lines[idx + 1]
        return ""

    return {
        "personal_number": after("개인번호"),
        "name": after("성명"),
        "job_category": after("직종"),
        "pay_step": after("급여연차"),
        "department": after("소속"),
        "hire_date": after("입사년월"),
    }


def page2_values_from_text(text: str) -> dict[str, Any]:
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    values: dict[str, Any] = {}
    for idx in range(0, len(lines) - 1, 2):
        key = lines[idx]
        value = lines[idx + 1]
        if key in {"개인번호", "이름"}:
            continue
        if re.search(r"[가-힣]", key):
            values[key] = parse_money(value) if re.search(r"\d", value) else value
    return values


def template_items_from_words(
    words: list[tuple[Any, ...]],
    label_grid: list[list[str]],
    y_centers: list[float],
    source: str,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row_index, row in enumerate(label_grid):
        y_center = y_centers[row_index]
        for col_index, label in enumerate(row):
            if not label:
                continue
            bbox = amount_bbox(col_index, y_center)
            raw_value = words_in_bbox(words, bbox)
            amount = parse_money(raw_value)
            if amount is None and not is_total_label(label):
                continue
            items.append(build_item(label, amount, raw_value, source, row_index, col_index))
    return items


def amount_bbox(col_index: int, y_center: float, pad_y: float = 7.0) -> list[float]:
    return [
        PDF_X_BOUNDS[col_index],
        y_center - pad_y,
        PDF_X_BOUNDS[col_index + 1],
        y_center + pad_y,
    ]


def words_in_bbox(words: list[tuple[Any, ...]], bbox: list[float]) -> str:
    x1, y1, x2, y2 = bbox
    values = []
    for word in words:
        wx1, wy1, wx2, wy2, text = word[:5]
        cx = (float(wx1) + float(wx2)) / 2
        cy = (float(wy1) + float(wy2)) / 2
        if x1 <= cx <= x2 and y1 <= cy <= y2:
            values.append(str(text))
    return " ".join(values)


def create_surya_predictor() -> Any:
    import os

    os.environ.setdefault("DISABLE_TQDM", "true")
    from surya.foundation import FoundationPredictor
    from surya.recognition import RecognitionPredictor

    foundation = FoundationPredictor()
    predictor = RecognitionPredictor(foundation)
    for item in (foundation, predictor):
        if hasattr(item, "disable_tqdm"):
            item.disable_tqdm = True
    return predictor


def ocr_image_bboxes(predictor: Any, image: Any, bboxes: list[list[int]]) -> list[dict[str, Any]]:
    from surya.common.surya.schema import TaskNames

    prediction = predictor(
        [image],
        task_names=[TaskNames.ocr_without_boxes],
        bboxes=[bboxes],
        math_mode=False,
        sort_lines=False,
    )[0]

    if len(prediction.text_lines) != len(bboxes):
        raise RuntimeError(f"Surya returned {len(prediction.text_lines)} OCR lines for {len(bboxes)} boxes")

    rows: list[dict[str, Any]] = []
    for bbox, line in zip(bboxes, prediction.text_lines):
        rows.append(
            {
                "bbox": bbox,
                "raw_text": clean_text(line.text or ""),
                "confidence": round(float(line.confidence or 0), 4),
            }
        )
    return rows


def parse_image(path: Path) -> dict[str, Any]:
    from PIL import Image

    start = time.time()
    image = Image.open(path).convert("RGB")
    width, height = image.size
    scale_x = width / PDF_WIDTH
    scale_y = height / PDF_HEIGHT
    bboxes: list[list[int]] = []
    keys: list[tuple[str, str, str | None]] = []
    for key, bbox in HEADER_BBOXES.items():
        bboxes.append(scale_bbox(bbox, scale_x, scale_y))
        keys.append(("header", key, None))
    for row_index, row in enumerate(PAYMENT_LABEL_GRID):
        for col_index, label in enumerate(row):
            if not label:
                continue
            bboxes.append(scale_bbox(amount_bbox(col_index, PDF_PAYMENT_Y_CENTERS[row_index], pad_y=8), scale_x, scale_y))
            keys.append(("earning", label, None))
    for row_index, row in enumerate(DEDUCTION_LABEL_GRID):
        for col_index, label in enumerate(row):
            if not label:
                continue
            bboxes.append(scale_bbox(amount_bbox(col_index, PDF_DEDUCTION_Y_CENTERS[row_index], pad_y=8), scale_x, scale_y))
            keys.append(("deduction", label, None))
    for key, bbox in SUPPLEMENTAL_IMAGE_BBOXES.items():
        bboxes.append(scale_bbox(bbox, scale_x, scale_y))
        keys.append(("supplemental", key, None))

    predictor = create_surya_predictor()
    ocr_rows = ocr_image_bboxes(predictor, image, bboxes)
    header: dict[str, str] = {}
    supplemental_raw: dict[str, str] = {}
    earnings: list[dict[str, Any]] = []
    deductions: list[dict[str, Any]] = []
    for key_info, ocr in zip(keys, ocr_rows):
        kind, label, _ = key_info
        raw_value = ocr["raw_text"]
        if kind == "header":
            header[label] = raw_value
            continue
        if kind == "supplemental":
            supplemental_raw[label] = raw_value
            continue
        amount = parse_money(raw_value)
        if amount is None and not is_total_label(label):
            continue
        item = build_item(label, amount, raw_value, kind, None, None)
        item["confidence"] = ocr.get("confidence")
        item["bbox"] = ocr.get("bbox")
        if kind == "earning":
            earnings.append(item)
        else:
            deductions.append(item)

    title = normalize_ocr_title(header.get("title") or path.stem)
    employee = {
        "personal_number": only_digits(header.get("personal_number")),
        "name": compact_text(header.get("name")),
        "job_category": compact_text(header.get("job_category")),
        "pay_step": clean_text(header.get("pay_step")),
        "department": compact_text(header.get("department")),
        "hire_date": normalize_date(header.get("hire_date")),
    }
    pay_date = normalize_date(header.get("pay_date"))
    statement = build_statement(path, "image-ocr", title, pay_date, employee, earnings, deductions, supplemental_from_image_raw(supplemental_raw))
    statement["elapsed_seconds"] = round(time.time() - start, 2)
    return statement


def scale_bbox(bbox: list[float], scale_x: float, scale_y: float) -> list[int]:
    return [
        int(round(bbox[0] * scale_x)),
        int(round(bbox[1] * scale_y)),
        int(round(bbox[2] * scale_x)),
        int(round(bbox[3] * scale_y)),
    ]


def only_digits(value: Any) -> str:
    return "".join(re.findall(r"\d+", str(value or "")))


def normalize_date(value: Any) -> str:
    text = clean_text(value)
    match = re.search(r"(\d{4})[-./년 ]+(\d{1,2})[-./월 ]+(\d{1,2})", text)
    if match:
        return f"{int(match.group(1)):04d}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    return text


def normalize_ocr_title(value: Any) -> str:
    text = compact_text(value)
    match = re.search(r"(\d{4}).*?(\d{1,2}).*?급여명세서", text)
    if match:
        return f"{int(match.group(1)):04d}년도 {int(match.group(2)):02d}월분 급여명세서"
    return clean_text(value)


def parse_csv_file(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        grid = [row for row in csv.reader(handle)]
    if not grid:
        return build_statement(path, "csv", path.stem, "", {}, [], [], {})

    header = [compact_text(value).lower() for value in grid[0]]
    if "section" not in header or "label" not in header:
        return parse_grid_like_excel(path, "csv-grid", grid)

    rows = [dict(zip(grid[0], row)) for row in grid[1:]]
    title = path.stem
    pay_date = ""
    employee: dict[str, str] = {}
    earnings: list[dict[str, Any]] = []
    deductions: list[dict[str, Any]] = []
    for row in rows:
        section = clean_text(row.get("section"))
        label = clean_text(row.get("label"))
        value = clean_text(row.get("value") or row.get("amount"))
        if section == "metadata":
            if label == "title":
                title = value
            elif label == "pay_date":
                pay_date = value
            else:
                employee[label] = value
        elif section in {"earning", "deduction"}:
            item = build_item(label, parse_money(value), value, section)
            (earnings if section == "earning" else deductions).append(item)
    return build_statement(path, "csv", title, pay_date, employee, earnings, deductions, {})


def parse_xlsx(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        grid = [[sheet.cell(row=row + 1, column=col + 1).value for col in range(sheet.max_column)] for row in range(sheet.max_row)]
    finally:
        workbook.close()
    return parse_grid_like_excel(path, "xlsx", grid)


def parse_grid_like_excel(path: Path, source_format: str, grid: list[list[Any]]) -> dict[str, Any]:
    def value(row: int, col: int) -> Any:
        if row >= len(grid) or col >= len(grid[row]):
            return ""
        return grid[row][col]

    class SheetAdapter:
        nrows = len(grid)
        ncols = max((len(row) for row in grid), default=0)

        @staticmethod
        def cell_value(row: int, col: int) -> Any:
            return value(row, col)

    sheet = SheetAdapter()
    title = first_cell_matching(sheet, "급여명세서") or path.stem
    employee = {
        "personal_number": clean_text(value_right_of_label(sheet, "개인번호")),
        "name": clean_text(value_right_of_label(sheet, "성 명") or value_right_of_label(sheet, "성명")),
        "job_category": clean_text(value_right_of_label(sheet, "직 종") or value_right_of_label(sheet, "직종")),
        "pay_step": clean_text(value_right_of_label(sheet, "급여연차")),
        "department": clean_text(value_right_of_label(sheet, "소 속") or value_right_of_label(sheet, "소속")),
        "hire_date": clean_text(value_right_of_label(sheet, "입사년월")),
    }
    return build_statement(
        path,
        source_format,
        title,
        excel_serial_date(value_right_of_label(sheet, "급여지급일")),
        employee,
        xls_items(sheet, EXCEL_PAYMENT_LABEL_ROWS, "earning", EXCEL_LABEL_COLS),
        xls_items(sheet, EXCEL_DEDUCTION_LABEL_ROWS, "deduction", EXCEL_DEDUCTION_AMOUNT_COLS),
        parse_page1_supplement_from_sheet(sheet),
    )


def build_statement(
    path: Path,
    source_format: str,
    title: str,
    pay_date: str,
    employee: dict[str, Any],
    earnings: list[dict[str, Any]],
    deductions: list[dict[str, Any]],
    supplemental: dict[str, Any],
) -> dict[str, Any]:
    totals = extract_totals(earnings)
    supplemental = attach_calculation_matches(merge_supplemental(supplemental), earnings, deductions)
    validation = validate_totals(totals)
    validation = add_supplemental_validation(validation, supplemental)
    score = quality_score(employee, totals, validation)
    return {
        "document_type": "payroll_statement",
        "source_file": str(path),
        "source_format": source_format,
        "generated_at": now_iso(),
        "metadata": {
            "title": clean_text(title),
            "pay_period": pay_period_from_title(title),
            "statement_kind": statement_kind_from_title(title),
            "pay_date": normalize_date(pay_date),
        },
        "employee": normalize_employee(employee),
        "earnings": sorted_items(earnings),
        "deductions": sorted_items(deductions),
        "supplemental": supplemental,
        "annual_leave": supplemental.get("work_basis", {}),
        "totals": totals,
        "validation": validation,
        "quality": {
            "score": score,
            "missing_core_fields": missing_core_fields(employee, totals),
            "earning_item_count": len(earnings),
            "deduction_item_count": len(deductions),
        },
    }


def attach_calculation_matches(
    supplemental: dict[str, Any],
    earnings: list[dict[str, Any]],
    deductions: list[dict[str, Any]],
) -> dict[str, Any]:
    pools = {
        "earning": {compact_text(item.get("label")): item.get("amount") for item in earnings},
        "deduction": {compact_text(item.get("label")): item.get("amount") for item in deductions},
    }
    details: list[dict[str, Any]] = []
    for detail in supplemental.get("calculation_details", []):
        enriched = dict(detail)
        matched = pools.get(str(detail.get("section")), {}).get(compact_text(detail.get("item")))
        enriched["matched_statement_amount"] = matched
        enriched["is_matched"] = matched == detail.get("amount") if matched is not None else False
        details.append(enriched)
    supplemental["calculation_details"] = details
    return supplemental


def add_supplemental_validation(validation: dict[str, Any], supplemental: dict[str, Any]) -> dict[str, Any]:
    mismatches = []
    warnings = []
    for detail in supplemental.get("calculation_details", []):
        matched = detail.get("matched_statement_amount")
        if matched is None:
            warnings.append({"rule": "calculation_detail_matches_statement_item", "item": detail.get("item")})
        elif not detail.get("is_matched"):
            mismatches.append(
                {
                    "rule": "calculation_detail_matches_statement_item",
                    "item": detail.get("item"),
                    "computed": detail.get("amount"),
                    "printed": matched,
                }
            )
    validation["supplemental_mismatches"] = mismatches
    validation["supplemental_warnings"] = warnings
    validation["is_valid"] = bool(validation.get("is_valid")) and not mismatches
    return validation


def sorted_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [item for item in items if item.get("amount") is not None]


def extract_totals(earnings: list[dict[str, Any]]) -> dict[str, int | None]:
    by_label = {compact_text(item["label"]): item.get("amount") for item in earnings}
    return {
        "gross_pay": by_label.get("급여총액"),
        "deduction_total": by_label.get("공제총액"),
        "net_pay": by_label.get("실지급액"),
    }


def validate_totals(totals: dict[str, int | None]) -> dict[str, Any]:
    gross = totals.get("gross_pay")
    deduction = totals.get("deduction_total")
    net = totals.get("net_pay")
    mismatches: list[dict[str, Any]] = []
    if None not in (gross, deduction, net):
        computed = int(gross) - int(deduction)
        if computed != int(net):
            mismatches.append({"rule": "gross_minus_deduction_equals_net", "computed": computed, "printed": net})
    return {
        "is_valid": not mismatches and None not in (gross, deduction, net),
        "mismatches": mismatches,
    }


def normalize_employee(employee: dict[str, Any]) -> dict[str, str]:
    return {
        "personal_number": only_digits(employee.get("personal_number")),
        "name": compact_text(employee.get("name")),
        "job_category": compact_text(employee.get("job_category")),
        "pay_step": clean_text(employee.get("pay_step")),
        "department": compact_text(employee.get("department")),
        "hire_date": normalize_date(employee.get("hire_date")),
    }


def missing_core_fields(employee: dict[str, Any], totals: dict[str, int | None]) -> list[str]:
    normalized = normalize_employee(employee)
    missing = [key for key in ["personal_number", "name"] if not normalized.get(key)]
    missing.extend(key for key, value in totals.items() if value is None)
    return missing


def quality_score(employee: dict[str, Any], totals: dict[str, int | None], validation: dict[str, Any]) -> float:
    checks = 6
    defects = len(missing_core_fields(employee, totals)) + len(validation.get("mismatches", []))
    return round(max(0.0, 1.0 - defects / checks), 4)


def render_html(statement: dict[str, Any], output_path: Path) -> None:
    def esc(value: Any) -> str:
        return html.escape("" if value is None else str(value))

    def won(value: Any) -> str:
        amount = format_amount(value)
        return f"{amount}원" if amount else "-"

    def is_total_item(item: dict[str, Any]) -> bool:
        return compact_text(item.get("label")) in {"급여총액", "공제총액", "실지급액"}

    def transaction_rows(items: list[dict[str, Any]], kind: str) -> str:
        filtered = [item for item in items if not is_total_item(item) and item.get("amount") not in (None, 0)]
        if not filtered:
            return '<div class="empty-state">표시할 항목 없음</div>'
        sign = "-" if kind == "deduction" else "+"
        groups = PAY_GROUPS if kind == "income" else DEDUCTION_GROUPS
        grouped: list[tuple[str, list[dict[str, Any]]]] = []
        used: set[int] = set()
        for group_name, labels in groups:
            label_set = {compact_text(label) for label in labels}
            group_items = [item for index, item in enumerate(filtered) if index not in used and compact_text(item.get("label")) in label_set]
            if group_items:
                for index, item in enumerate(filtered):
                    if item in group_items:
                        used.add(index)
                grouped.append((group_name, group_items))
        leftovers = [item for index, item in enumerate(filtered) if index not in used]
        if leftovers:
            grouped.append(("기타", leftovers))

        parts = []
        for group_name, group_items in grouped:
            group_total = sum(int(item.get("amount") or 0) for item in group_items)
            body = "".join(
                f"""<article class="transaction {kind}">
          <div>
            <strong>{esc(item.get('label'))}</strong>
          </div>
          <b>{esc(sign)}{esc(won(item.get('amount')))}</b>
        </article>"""
                for item in group_items
            )
            parts.append(
                f"""<details class="paylab-group {kind}" open>
        <summary><span>{esc(group_name)}</span><b>{esc(won(group_total))}</b></summary>
        <div>{body}</div>
      </details>"""
            )
        return "".join(parts)

    def source_preview() -> str:
        asset = statement.get("source_asset")
        if not asset:
            return ""
        suffix = Path(str(asset)).suffix.lower()
        if suffix in IMAGE_EXTENSIONS:
            body = f'<img src="{esc(asset)}" alt="원본 급여명세서">'
        elif suffix == ".pdf":
            body = f'<embed src="{esc(asset)}" type="application/pdf">'
        else:
            body = f'<a class="source-file" href="{esc(asset)}">{esc(asset)}</a>'
        return f"""<details class="source-panel">
    <summary>원본 보기</summary>
    <div class="source-body">{body}</div>
  </details>"""

    def source_note(item: dict[str, Any]) -> str:
        confidence = item.get("confidence")
        raw_value = item.get("raw_value")
        if confidence is not None:
            return f"OCR {float(confidence):.2f} · 원문 {raw_value}"
        return f"원문 {raw_value}"

    def admin_item_rows(items: list[dict[str, Any]], kind: str) -> str:
        filtered = [item for item in items if item.get("amount") not in (None, 0)]
        if not filtered:
            return '<div class="empty-state">표시할 파서 항목 없음</div>'
        sign = "-" if kind == "deduction" else "+"
        return "".join(
            f"""<article class="admin-item">
        <div>
          <strong>{esc(item.get('label'))}</strong>
          <span>{esc(source_note(item))}</span>
        </div>
        <b>{esc(sign)}{esc(won(item.get('amount')))}</b>
      </article>"""
            for item in filtered
        )

    def metadata_item(label: str, value: Any) -> str:
        return f'<div class="meta-item"><span>{esc(label)}</span><strong>{esc(value or "-")}</strong></div>'

    def total_card(label: str, value: Any, class_name: str, money: bool = True) -> str:
        display_value = won(value) if money else (value or "-")
        return f'<div class="total-card {class_name}"><span>{esc(label)}</span><strong>{esc(display_value)}</strong></div>'

    def basis_card(label: str, value: Any, money: bool = False) -> str:
        display_value = won(value) if money else (value if value not in (None, "") else "-")
        return f'<div class="basis-card"><span>{esc(label)}</span><strong>{esc(display_value)}</strong></div>'

    def admin_card(label: str, value: Any) -> str:
        return f'<div class="admin-card"><span>{esc(label)}</span><strong>{esc(value if value not in (None, "") else "-")}</strong></div>'

    def calculation_rows() -> str:
        details = supplemental.get("calculation_details", [])
        if not details:
            return '<div class="empty-state">표시할 계산근거 없음</div>'
        rows = []
        for detail in details:
            section_label = "공제" if detail.get("section") == "deduction" else "지급"
            match_text = "명세서와 일치" if detail.get("is_matched") else "대조 필요"
            rows.append(
                f"""<article class="calc-row">
        <div>
          <span>{esc(section_label)}</span>
          <strong>{esc(detail.get('item'))}</strong>
          <em>{esc(detail.get('formula') or '-')}</em>
        </div>
        <div>
          <b>{esc(won(detail.get('amount')))}</b>
          <small>{esc(match_text)}</small>
        </div>
      </article>"""
            )
        return "".join(rows)

    def contact_rows() -> str:
        contacts = supplemental.get("contacts", [])
        if not contacts:
            return '<div class="empty-state">표시할 문의처 없음</div>'
        return "".join(
            f"""<article class="contact-row">
        <div>
          <strong>{esc(contact.get('topic') or '급여문의')}</strong>
          <span>{esc(' · '.join(part for part in [contact.get('team'), contact.get('category')] if part))}</span>
        </div>
        <b>{esc(contact.get('extension'))}</b>
      </article>"""
            for contact in contacts
        )

    def message_rows() -> str:
        messages = supplemental.get("messages", {})
        entries = []
        if messages.get("family_allowance_target"):
            entries.append(("가족수당 지급대상", messages["family_allowance_target"]))
        if messages.get("payroll_account_change_instruction"):
            entries.append(("급여계좌 변경", messages["payroll_account_change_instruction"]))
        if not entries:
            return ""
        return "".join(
            f'<div class="message-row"><span>{esc(label)}</span><strong>{esc(value)}</strong></div>'
            for label, value in entries
        )

    def work_metric_rows() -> str:
        metrics = [item for item in supplemental.get("work_metrics", []) if item.get("amount") not in (None, "")]
        if not metrics:
            return ""
        return '<div class="metric-strip">' + "".join(
            f'<div><span>{esc(item.get("label"))}</span><strong>{esc(format_amount(item.get("amount")))}</strong></div>'
            for item in metrics[:8]
        ) + "</div>"

    totals = statement["totals"]
    employee = statement["employee"]
    metadata = statement["metadata"]
    validation = statement["validation"]
    supplemental = merge_supplemental(statement.get("supplemental", {}))
    membership = supplemental.get("membership", {})
    work_basis = supplemental.get("work_basis", {})
    is_valid = bool(validation.get("is_valid"))
    status_text = "검증 완료" if is_valid else "확인 필요"
    status_class = "ok" if is_valid else "bad"
    source_format = statement.get("source_format", "")
    quality_score = statement.get("quality", {}).get("score")
    quality_text = f"{float(quality_score) * 100:.1f}%" if isinstance(quality_score, int | float) else "-"
    elapsed_seconds = statement.get("elapsed_seconds")
    elapsed_text = f"{elapsed_seconds}s" if elapsed_seconds not in (None, "") else "-"
    gross_pay = int(totals.get("gross_pay") or 0)
    deduction_total = int(totals.get("deduction_total") or 0)
    net_ratio = round(((gross_pay - deduction_total) / gross_pay * 100), 1) if gross_pay else 0
    deduction_ratio = round(max(0, 100 - net_ratio), 1) if gross_pay else 0

    output_path.write_text(
        f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{esc(metadata['title'])}</title>
  <style>
    :root {{
      --bg: #08080c;
      --surface: #1a1a28;
      --surface-soft: #222234;
      --ink: #eaeaf2;
      --muted: #8888a8;
      --quiet: #5a5a78;
      --line: #2a2a3e;
      --line-2: #363650;
      --brand: #00e5a0;
      --brand-2: #5c9aff;
      --income: #00e5a0;
      --deduction: #ff5c72;
      --net: #ffd740;
      --purple: #a78bfa;
      --warn: #ffb86c;
      --mono: "Consolas", "IBM Plex Mono", monospace;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      min-height: 100vh;
      background: var(--bg);
      color: var(--ink);
      font-family: "Noto Sans KR", "Malgun Gothic", Arial, sans-serif;
      letter-spacing: 0;
    }}
    .bank-shell {{
      width: min(560px, calc(100% - 32px));
      margin: 0 auto;
      padding: 20px 0 88px;
    }}
    .topbar {{
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
      margin-bottom: 16px;
    }}
    .brand {{ display:flex; align-items:center; gap:10px; min-width:0; }}
    .brand-mark {{ width:32px; height:32px; border-radius:9px; display:flex; align-items:center; justify-content:center; color:#08080c; font-weight:900; background:linear-gradient(135deg,var(--income),var(--brand-2)); }}
    .brand-name {{ font-weight:800; font-size:15px; }}
    .brand-sub {{ margin-top:2px; color:var(--quiet); font-size:11px; overflow-wrap:anywhere; }}
    .eyebrow {{
      margin: 0 0 5px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 700;
    }}
    h1 {{
      margin: 0;
      font-size: clamp(20px, 2.4vw, 30px);
      line-height: 1.22;
    }}
    h2 {{
      margin: 0;
      font-size: 18px;
      line-height: 1.3;
    }}
    .status {{
      flex: 0 0 auto;
      border: 1px solid var(--line);
      border-radius: 999px;
      padding: 7px 11px;
      background: var(--surface);
      font-size: 13px;
      font-weight: 800;
    }}
    .status.ok {{ color: var(--income); }}
    .status.bad {{ color: var(--deduction); }}
    .balance-card {{
      text-align: center;
      border-radius: 16px;
      padding: 22px 0 8px;
      background: transparent;
      color: var(--ink);
      border: 0;
    }}
    .balance-card span {{
      display: block;
      color: var(--quiet);
      font-size: 11px;
      font-weight: 800;
      letter-spacing: 1.5px;
      text-transform: uppercase;
    }}
    .balance-amount {{
      margin-top: 8px;
      font-family: var(--mono);
      font-size: clamp(38px, 9vw, 52px);
      line-height: 1;
      font-weight: 800;
      color: var(--net);
      letter-spacing: -2px;
    }}
    .hero-sub {{ margin-top:7px; color:var(--quiet); font-size:11px; }}
    .balance-subgrid {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 0;
      margin-top: 18px;
      border: 1px solid var(--line);
      border-radius: 16px;
      background: var(--surface);
      overflow: hidden;
    }}
    .balance-subgrid div {{
      padding: 13px 8px;
      min-width: 0;
      position: relative;
    }}
    .balance-subgrid div:not(:last-child)::after {{
      content:"";
      position:absolute;
      right:0;
      top:18%;
      bottom:18%;
      width:1px;
      background:var(--line);
    }}
    .balance-subgrid strong {{
      display: block;
      margin-top: 4px;
      overflow-wrap: anywhere;
      font-size: 15px;
      font-family: var(--mono);
    }}
    .ratio-bar {{ margin:14px auto 0; max-width:300px; height:6px; background:var(--surface); border-radius:3px; overflow:hidden; display:flex; }}
    .rb-e {{ background:var(--income); }}
    .rb-d {{ background:var(--deduction); }}
    .ratio-labels {{ display:flex; justify-content:space-between; max-width:300px; margin:6px auto 0; font-size:10px; color:var(--quiet); }}
    .meta-strip {{
      display: grid;
      grid-template-columns: repeat(4, minmax(0, 1fr));
      gap: 8px;
      margin: 12px 0;
    }}
    .total-strip {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 8px;
      margin: 12px 0;
    }}
    .meta-item, .total-card, .panel, .source-panel, .detail-panel, .admin-panel {{
      border: 1px solid var(--line);
      border-radius: 16px;
      background: var(--surface);
    }}
    .meta-item, .total-card {{
      padding: 12px;
      min-width: 0;
    }}
    .meta-item span, .total-card span {{
      display: block;
      color: var(--muted);
      font-size: 12px;
      font-weight: 800;
    }}
    .meta-item strong, .total-card strong {{
      display: block;
      margin-top: 5px;
      font-size: 17px;
      overflow-wrap: anywhere;
    }}
    .total-card.net strong {{ color: var(--net); }}
    .total-card.in strong {{ color: var(--income); }}
    .total-card.out strong {{ color: var(--deduction); }}
    .content-grid {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      gap: 12px;
      align-items: start;
    }}
    .insight-grid {{
      display: grid;
      grid-template-columns: minmax(0, 1.08fr) minmax(0, .92fr);
      gap: 12px;
      align-items: start;
      margin: 12px 0;
    }}
    .panel {{
      overflow: hidden;
    }}
    .panel header {{
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      padding: 14px 16px;
      border-bottom: 1px solid var(--line);
      background: var(--surface-soft);
    }}
    .panel header b {{
      white-space: nowrap;
      font-size: 15px;
    }}
    .transaction {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 12px;
      align-items: center;
      padding: 13px 16px;
      border-bottom: 1px solid var(--line);
    }}
    .transaction:last-child {{ border-bottom: 0; }}
    .transaction strong {{
      display: block;
      line-height: 1.35;
      overflow-wrap: anywhere;
    }}
    .transaction span {{
      display: block;
      margin-top: 3px;
      color: var(--muted);
      font-size: 12px;
      overflow-wrap: anywhere;
    }}
    .transaction b {{
      text-align: right;
      white-space: nowrap;
      font-size: 15px;
    }}
    .transaction.income b {{ color: var(--income); }}
    .transaction.deduction b {{ color: var(--deduction); }}
    .paylab-group {{ border-bottom: 1px solid var(--line); }}
    .paylab-group:last-child {{ border-bottom: 0; }}
    .paylab-group summary {{
      cursor: pointer;
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 10px;
      list-style: none;
      padding: 10px 16px 10px 24px;
      color: var(--muted);
      font-size: 13px;
      font-weight: 800;
    }}
    .paylab-group summary::-webkit-details-marker {{ display:none; }}
    .paylab-group summary span {{ display:flex; align-items:center; gap:8px; }}
    .paylab-group summary span::before {{
      content:"";
      width:5px;
      height:5px;
      border-radius:50%;
      background: var(--income);
    }}
    .paylab-group.deduction summary span::before {{ background: var(--deduction); }}
    .paylab-group summary b {{ font-family:var(--mono); color:var(--ink); font-size:12px; }}
    .paylab-group .transaction {{ padding-left: 42px; }}
    .basis-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      padding: 14px 16px;
    }}
    .basis-card {{
      min-width: 0;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px;
      background: var(--surface-soft);
    }}
    .basis-card span, .message-row span {{
      display: block;
      color: var(--muted);
      font-size: 12px;
      font-weight: 800;
    }}
    .basis-card strong, .message-row strong {{
      display: block;
      margin-top: 4px;
      overflow-wrap: anywhere;
    }}
    .calc-row, .contact-row {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 12px;
      align-items: center;
      padding: 13px 16px;
      border-bottom: 1px solid var(--line);
    }}
    .calc-row:last-child, .contact-row:last-child {{ border-bottom: 0; }}
    .calc-row span {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 800;
    }}
    .calc-row strong, .contact-row strong {{
      display: block;
      margin-top: 3px;
      overflow-wrap: anywhere;
    }}
    .calc-row em {{
      display: block;
      margin-top: 3px;
      color: var(--muted);
      font-size: 12px;
      font-style: normal;
      overflow-wrap: anywhere;
    }}
    .calc-row b {{
      display: block;
      text-align: right;
      white-space: nowrap;
    }}
    .calc-row small {{
      display: block;
      margin-top: 3px;
      color: var(--income);
      text-align: right;
      font-weight: 800;
    }}
    .contact-row span {{
      display: block;
      margin-top: 3px;
      color: var(--muted);
      font-size: 12px;
    }}
    .contact-row b {{
      color: var(--brand-2);
      font-size: 18px;
    }}
    .message-row {{
      border-top: 1px solid var(--line);
      padding: 12px 16px;
      background: var(--surface-soft);
    }}
    .detail-panel, .admin-panel {{
      margin-top: 12px;
      overflow: hidden;
    }}
    .detail-panel > summary, .admin-panel > summary {{
      cursor: pointer;
      display: flex;
      justify-content: space-between;
      align-items: center;
      gap: 12px;
      list-style: none;
      padding: 14px 16px;
      font-weight: 900;
      background: var(--surface-soft);
    }}
    .detail-panel > summary::-webkit-details-marker,
    .admin-panel > summary::-webkit-details-marker {{
      display: none;
    }}
    .detail-panel > summary span,
    .admin-panel > summary span {{
      color: var(--muted);
      font-size: 12px;
      font-weight: 800;
    }}
    .detail-body, .admin-body {{
      border-top: 1px solid var(--line);
    }}
    .detail-note, .admin-note {{
      margin: 0;
      padding: 12px 16px;
      color: var(--muted);
      border-bottom: 1px solid var(--line);
      font-size: 13px;
      line-height: 1.5;
    }}
    .detail-subhead {{
      margin: 0;
      padding: 12px 16px;
      color: var(--muted);
      border-top: 1px solid var(--line);
      border-bottom: 1px solid var(--line);
      font-size: 13px;
      font-weight: 900;
    }}
    .admin-grid {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      padding: 14px 16px;
    }}
    .admin-card {{
      min-width: 0;
      border: 1px solid var(--line);
      border-radius: 10px;
      padding: 10px;
      background: var(--surface-soft);
    }}
    .admin-card span {{
      display: block;
      color: var(--muted);
      font-size: 11px;
      font-weight: 900;
    }}
    .admin-card strong {{
      display: block;
      margin-top: 5px;
      overflow-wrap: anywhere;
      font-family: var(--mono);
      font-size: 12px;
    }}
    .admin-links {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      padding: 0 16px 14px;
    }}
    .admin-link {{
      color: var(--ink);
      border: 1px solid var(--line-2);
      border-radius: 999px;
      padding: 7px 10px;
      text-decoration: none;
      font-size: 12px;
      font-weight: 900;
      background: var(--surface-soft);
    }}
    .admin-split {{
      display: grid;
      grid-template-columns: 1fr 1fr;
      border-top: 1px solid var(--line);
    }}
    .admin-split section:first-child {{
      border-right: 1px solid var(--line);
    }}
    .admin-split h3 {{
      margin: 0;
      padding: 12px 16px;
      color: var(--muted);
      font-size: 13px;
      border-bottom: 1px solid var(--line);
    }}
    .admin-item {{
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 12px;
      padding: 11px 16px;
      border-bottom: 1px solid var(--line);
      align-items: center;
    }}
    .admin-item:last-child {{
      border-bottom: 0;
    }}
    .admin-item strong, .admin-item span {{
      display: block;
      overflow-wrap: anywhere;
    }}
    .admin-item span {{
      margin-top: 3px;
      color: var(--muted);
      font-size: 11px;
    }}
    .admin-item b {{
      white-space: nowrap;
      font-size: 12px;
      font-family: var(--mono);
    }}
    .metric-strip {{
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 8px;
      padding: 0 16px 14px;
    }}
    .metric-strip div {{
      border: 1px solid var(--line);
      border-radius: 10px;
      background: var(--surface-soft);
      padding: 10px;
      min-width: 0;
    }}
    .metric-strip span {{
      display:block;
      color:var(--muted);
      font-size:11px;
      font-weight:800;
      overflow-wrap:anywhere;
    }}
    .metric-strip strong {{
      display:block;
      margin-top:4px;
      font-family:var(--mono);
    }}
    .reconcile {{
      margin-top: 12px;
      border: 1px solid var(--line);
      border-radius: 16px;
      background: var(--surface);
      padding: 14px 16px;
      display: flex;
      justify-content: space-between;
      gap: 12px;
      align-items: center;
    }}
    .reconcile span {{
      color: var(--muted);
      font-size: 13px;
      font-weight: 800;
    }}
    .reconcile strong {{
      color: var(--income);
      font-size: 17px;
    }}
    .reconcile.bad strong {{ color: var(--deduction); }}
    .source-panel {{
      margin-top: 12px;
      overflow: hidden;
    }}
    .source-panel summary {{
      cursor: pointer;
      padding: 13px 16px;
      font-weight: 800;
      list-style-position: inside;
    }}
    .source-body {{
      border-top: 1px solid var(--line);
      padding: 12px;
      background: var(--surface-soft);
    }}
    .source-body img, .source-body embed {{
      width: 100%;
      max-height: 760px;
      object-fit: contain;
      border: 1px solid var(--line);
      border-radius: 10px;
      background: white;
    }}
    .source-body embed {{ min-height: 620px; }}
    .source-file {{
      color: var(--brand-2);
      font-weight: 800;
    }}
    .empty-state {{
      padding: 16px;
      color: var(--muted);
    }}
    @media (max-width: 900px) {{
      .content-grid, .insight-grid, .admin-split {{ grid-template-columns: 1fr; }}
      .admin-split section:first-child {{ border-right: 0; border-bottom: 1px solid var(--line); }}
      .meta-strip, .total-strip {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
    }}
    @media (max-width: 560px) {{
      .bank-shell {{ width: min(100% - 18px, 1120px); padding-top: 10px; }}
      .topbar {{ align-items: stretch; flex-direction: column; }}
      .status {{ width: fit-content; }}
      .balance-card {{ padding: 18px; }}
      .balance-subgrid {{ grid-template-columns: 1fr; }}
      .meta-strip, .total-strip {{ grid-template-columns: 1fr; }}
      .admin-grid {{ grid-template-columns: 1fr; }}
      .basis-grid {{ grid-template-columns: 1fr; }}
      .metric-strip {{ grid-template-columns: 1fr; }}
      .transaction {{ grid-template-columns: 1fr; }}
      .transaction b {{ text-align: left; }}
      .calc-row, .contact-row, .admin-item {{ grid-template-columns: 1fr; }}
      .calc-row b, .calc-row small {{ text-align: left; }}
      .reconcile {{ align-items: flex-start; flex-direction: column; }}
    }}
  </style>
</head>
<body>
<main class="shell bank-shell">
  <header class="topbar">
    <div class="brand">
      <div class="brand-mark">P</div>
      <div>
        <div class="brand-name">PayLab</div>
        <div class="brand-sub">{esc(metadata['title'])}</div>
      </div>
    </div>
    <span class="status {status_class}">{esc(status_text)}</span>
  </header>

  <section class="balance-card" aria-label="실수령액">
    <span>실수령액</span>
    <div class="balance-amount">{esc(won(totals.get('net_pay')))}</div>
    <div class="hero-sub">{esc(employee.get('name') or '-')} · {esc(employee.get('job_category') or '-')} · {esc(employee.get('pay_step') or '-')}</div>
    <div class="ratio-bar"><div class="rb-e" style="width:{net_ratio}%"></div><div class="rb-d" style="width:{deduction_ratio}%"></div></div>
    <div class="ratio-labels"><span style="color:var(--income)">실수령 {net_ratio}%</span><span style="color:var(--deduction)">공제 {deduction_ratio}%</span></div>
    <div class="balance-subgrid">
      <div><span>급여총액</span><strong>{esc(won(totals.get('gross_pay')))}</strong></div>
      <div><span>공제총액</span><strong>{esc(won(totals.get('deduction_total')))}</strong></div>
      <div><span>지급일</span><strong>{esc(metadata.get('pay_date') or "-")}</strong></div>
    </div>
  </section>

  <section class="meta-strip" aria-label="직원 정보">
    {metadata_item("성명", employee.get('name'))}
    {metadata_item("개인번호", employee.get('personal_number'))}
    {metadata_item("소속", employee.get('department'))}
    {metadata_item("급여월", metadata.get('pay_period'))}
  </section>

  <section class="total-strip" aria-label="합계">
    {total_card("지급 합계", totals.get('gross_pay'), "in")}
    {total_card("공제 합계", totals.get('deduction_total'), "out")}
    {total_card("최종 입금", totals.get('net_pay'), "net")}
  </section>

  <section class="content-grid">
    <section class="panel">
      <header><h2>지급내역</h2><b>{esc(won(totals.get('gross_pay')))}</b></header>
      {transaction_rows(statement['earnings'], "income")}
    </section>

    <section class="panel">
      <header><h2>공제내역</h2><b>{esc(won(totals.get('deduction_total')))}</b></header>
      {transaction_rows(statement['deductions'], "deduction")}
    </section>
  </section>

  <section class="reconcile {status_class}">
    <div>
      <span>합계 검증</span>
      <h2>급여총액 - 공제총액 = 실수령액</h2>
    </div>
    <strong>{esc(status_text)}</strong>
  </section>

  <section class="panel" style="margin-top:12px">
    <header><h2>문의처</h2><b>{esc(len(supplemental.get('contacts', [])))}건</b></header>
    {contact_rows()}
  </section>

  <details class="detail-panel">
    <summary>
      <div>
        <strong>세부 계산 기준</strong>
        <span>원본 2페이지/보조 정보</span>
      </div>
      <b>{esc(work_basis.get('work_days') or '-')}일 · {esc(len(supplemental.get('calculation_details', [])))}건</b>
    </summary>
    <div class="detail-body">
      <p class="detail-note">실수령액 확인에는 보통 필요하지 않은 산식, 근로일수, 연차평균임금, 공제 회원번호를 따로 모았습니다.</p>
      <div class="basis-grid">
        {basis_card("근로일수", work_basis.get('work_days'))}
        {basis_card("전월통상임금", work_basis.get('previous_month_regular_wage'), True)}
        {basis_card("연차평균임금", work_basis.get('annual_average_wage'), True)}
        {basis_card("교원공제회원번호", membership.get('teacher_mutual_aid_member_number'))}
      </div>
      {work_metric_rows()}
      {message_rows()}
      <h3 class="detail-subhead">계산근거</h3>
      {calculation_rows()}
    </div>
  </details>

  <details class="admin-panel">
    <summary>
      <div>
        <strong>관리자/검수 데이터</strong>
        <span>서버 저장, OCR 검수, 파서 개선용</span>
      </div>
      <b>{esc(source_format or '-')} · {esc(quality_text)}</b>
    </summary>
    <div class="admin-body">
      <p class="admin-note">이 영역은 사용자 급여 확인 화면과 분리된 내부 데이터입니다. API 저장, 파서 회귀 테스트, OCR 오류 분석에만 사용합니다.</p>
      <div class="admin-grid">
        {admin_card("원본형식", source_format)}
        {admin_card("OCR 품질", quality_text)}
        {admin_card("검증상태", status_text)}
        {admin_card("처리시간", elapsed_text)}
        {admin_card("경고", len(validation.get('warnings', [])))}
        {admin_card("불일치", len(validation.get('mismatches', [])))}
        {admin_card("원본파일", statement.get('source_file'))}
        {admin_card("문서유형", statement.get('document_type'))}
      </div>
      <div class="admin-links">
        <a class="admin-link" href="payroll.json">JSON</a>
        <a class="admin-link" href="payroll.csv">CSV</a>
        <a class="admin-link" href="payroll.xlsx">XLSX</a>
      </div>
      <div class="admin-split">
        <section>
          <h3>지급 항목 원문</h3>
          {admin_item_rows(statement['earnings'], "income")}
        </section>
        <section>
          <h3>공제 항목 원문</h3>
          {admin_item_rows(statement['deductions'], "deduction")}
        </section>
      </div>
      {source_preview()}
    </div>
  </details>
</main>
</body>
</html>
""",
        encoding="utf-8",
    )


def statement_rows(statement: dict[str, Any]) -> list[list[Any]]:
    rows: list[list[Any]] = [["section", "key", "label", "amount", "raw_value"]]
    for key, value in statement.get("metadata", {}).items():
        rows.append(["metadata", key, key, "", value])
    for key, value in statement.get("employee", {}).items():
        rows.append(["employee", key, key, "", value])
    supplemental = statement.get("supplemental", {})
    for key, value in supplemental.get("membership", {}).items():
        rows.append(["supplemental.membership", key, key, "", value])
    for key, value in supplemental.get("work_basis", {}).items():
        rows.append(["supplemental.work_basis", key, key, value if isinstance(value, int | float) else "", value])
    for metric in supplemental.get("work_metrics", []):
        rows.append(["supplemental.work_metric", compact_text(metric.get("label")), metric.get("label"), metric.get("amount"), metric.get("raw_value")])
    for key, value in supplemental.get("messages", {}).items():
        rows.append(["supplemental.messages", key, key, "", value])
    for contact in supplemental.get("contacts", []):
        label = " / ".join(part for part in [contact.get("topic"), contact.get("team"), contact.get("category")] if part)
        rows.append(["supplemental.contact", contact.get("extension", ""), label, "", contact.get("extension", "")])
    for detail in supplemental.get("calculation_details", []):
        rows.append(
            [
                "supplemental.calculation",
                detail.get("item"),
                detail.get("formula"),
                detail.get("amount"),
                f"matched={detail.get('matched_statement_amount')}; ok={detail.get('is_matched')}",
            ]
        )
    for key, value in statement.get("totals", {}).items():
        rows.append(["total", key, key, value, value])
    rows.append(["validation", "is_valid", "is_valid", "", statement.get("validation", {}).get("is_valid")])
    for section in ["earnings", "deductions"]:
        for item in statement.get(section, []):
            rows.append(
                [
                    "earning" if section == "earnings" else "deduction",
                    compact_text(item.get("label")),
                    item.get("label"),
                    item.get("amount"),
                    item.get("raw_value"),
                ]
            )
    return rows


def write_statement_csv(statement: dict[str, Any], output_path: Path) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        csv.writer(handle).writerows(statement_rows(statement))


def write_statement_xlsx(statement: dict[str, Any], output_path: Path) -> None:
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "payroll"
    for row in statement_rows(statement):
        sheet.append(row)
    workbook.save(output_path)


def format_amount(value: Any) -> str:
    if value is None:
        return ""
    try:
        return f"{int(value):,}"
    except (TypeError, ValueError):
        return str(value)


def parse_payroll_file(path: Path, output_dir: Path = DEFAULT_OUTPUT) -> dict[str, Any]:
    start = time.time()
    output_dir.mkdir(parents=True, exist_ok=True)
    kind = detect_payroll_file_kind(path)
    if kind == "xls-binary":
        statement = parse_xls(path)
    elif kind == "pdf":
        statement = parse_pdf(path)
    elif kind == "image":
        statement = parse_image(path)
    elif kind == "csv":
        statement = parse_csv_file(path)
    elif kind == "xlsx":
        statement = parse_xlsx(path)
    else:
        raise ValueError(f"Unsupported payroll file type: {path.suffix or 'unknown'}")
    statement.setdefault("elapsed_seconds", round(time.time() - start, 2))

    asset = copy_source_asset(path, output_dir)
    statement["source_asset"] = asset
    json_path = output_dir / "payroll.json"
    html_path = output_dir / "index.html"
    csv_path = output_dir / "payroll.csv"
    xlsx_path = output_dir / "payroll.xlsx"
    json_path.write_text(json.dumps(statement, ensure_ascii=False, indent=2), encoding="utf-8")
    render_html(statement, html_path)
    write_statement_csv(statement, csv_path)
    write_statement_xlsx(statement, xlsx_path)
    return {
        "json": json_path,
        "html": html_path,
        "csv": csv_path,
        "xlsx": xlsx_path,
        "employee_name": statement["employee"]["name"],
        "net_pay": statement["totals"]["net_pay"],
        "quality_score": statement["quality"]["score"],
        "elapsed_seconds": statement["elapsed_seconds"],
    }


def copy_source_asset(path: Path, output_dir: Path) -> str:
    target = output_dir / path.name
    try:
        if path.resolve() != target.resolve():
            shutil.copy2(path, target)
    except FileNotFoundError:
        return ""
    return target.name


def main() -> None:
    parser = argparse.ArgumentParser(description="Parse payroll statements from PDF, image, CSV, XLS, or XLSX.")
    parser.add_argument("files", nargs="+", type=Path)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    results = []
    for file_path in args.files:
        out_dir = args.output_dir / file_path.stem
        result = parse_payroll_file(file_path, out_dir)
        results.append({key: str(value) for key, value in result.items()})
    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "batch_summary.json").write_text(
        json.dumps(results, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(results, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
