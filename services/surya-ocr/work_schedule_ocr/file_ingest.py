from __future__ import annotations

import csv
import datetime as dt
from html.parser import HTMLParser
import json
import shutil
import time
from pathlib import Path
from typing import Any, Iterable, Sequence

from PIL import Image

from format_exports import REQUIRED, SUMMARY_FIELD_NAMES, SUMMARY_KEYS, STRUCTURED_HEADERS
from parse_batch import (
    DEFAULT_CODEBOOK,
    DEFAULT_DEPARTMENT_CODEBOOK_DIR,
    DEFAULT_SCHEDULE_PROFILE_DIR,
    codebook_paths_for_image,
    department_id_from_image,
    schedule_profile_path_for_image,
)
from schedule_quality import active_days_from_title, blocking_mismatches, parse_int, validate_column_profile, validate_employee_summary


IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
CSV_EXTENSIONS = {".csv", ".tsv"}
XLSX_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def detect_file_kind(path: Path) -> str:
    suffix = path.suffix.lower()
    prefix = Path(path).read_bytes()[:4096]
    stripped = prefix.lstrip().lower()
    if prefix.startswith(b"\xff\xd8\xff") or prefix.startswith(b"\x89PNG\r\n\x1a\n"):
        return "image"
    if prefix.startswith(b"%PDF"):
        return "pdf"
    if prefix.startswith(b"PK\x03\x04"):
        return "xlsx"
    if prefix.startswith(b"\xd0\xcf\x11\xe0"):
        return "xls-binary"
    if suffix in CSV_EXTENSIONS or looks_like_csv(prefix):
        return "csv"
    if suffix in IMAGE_EXTENSIONS:
        return "image"
    if suffix in XLSX_EXTENSIONS:
        return "xlsx"
    if suffix == ".xls" and (b"<table" in stripped or b"<html" in stripped):
        return "xls-html"
    if suffix == ".pdf":
        return "pdf"
    return "unknown"


def looks_like_csv(prefix: bytes) -> bool:
    try:
        text = prefix.decode("utf-8-sig")
    except UnicodeDecodeError:
        return False
    first_line = text.splitlines()[0] if text.splitlines() else ""
    return "," in first_line and any(header in first_line for header in ["day_1", "title", "name"])


def read_csv_records(path: Path) -> list[dict[str, str]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        return [
            {str(key): clean_value(value) for key, value in row.items() if key is not None}
            for row in reader
        ]


class TableHtmlParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.in_cell = False
        self.current_cell: list[str] = []
        self.current_row: list[str] = []
        self.rows: list[list[str]] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag.lower() in {"td", "th"}:
            self.in_cell = True
            self.current_cell = []

    def handle_data(self, data: str) -> None:
        if self.in_cell:
            self.current_cell.append(data)

    def handle_endtag(self, tag: str) -> None:
        tag = tag.lower()
        if tag in {"td", "th"} and self.in_cell:
            self.current_row.append(clean_value("".join(self.current_cell)))
            self.current_cell = []
            self.in_cell = False
        elif tag == "tr":
            if self.current_row:
                self.rows.append(self.current_row)
            self.current_row = []


def read_xls_html_records(path: Path) -> list[dict[str, str]]:
    parser = TableHtmlParser()
    parser.feed(Path(path).read_text(encoding="utf-8"))
    if not parser.rows:
        return []
    headers = parser.rows[0]
    return [
        {
            headers[index]: clean_value(value)
            for index, value in enumerate(row)
            if index < len(headers)
        }
        for row in parser.rows[1:]
    ]


def read_xlsx_records(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse xlsx files") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [clean_value(value) for value in rows[0]]
        records: list[dict[str, str]] = []
        for row in rows[1:]:
            records.append(
                {
                    headers[index]: clean_value(value)
                    for index, value in enumerate(row)
                    if index < len(headers) and headers[index]
                }
            )
        return records
    finally:
        workbook.close()


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n").strip()
    if text.endswith(".0") and text[:-2].lstrip("-").isdigit():
        return text[:-2]
    return text


def cell(value: Any, source_flag: str, bbox: list[int] | None = None) -> dict[str, Any]:
    text = clean_value(value)
    return {
        "value": text,
        "text": text,
        "raw_text": text,
        "confidence": 1.0,
        "bbox": bbox or [],
        "flags": [source_flag],
        "style": {},
        "candidates": [],
    }


def load_codebook_and_profile(input_path: Path, department_id: str | None):
    from parse_schedule import load_codebook, load_schedule_profile

    resolved_department = department_id or department_id_from_image(input_path)
    if resolved_department:
        department_codebook = DEFAULT_DEPARTMENT_CODEBOOK_DIR / f"{resolved_department}.json"
        codebook_paths = [DEFAULT_CODEBOOK]
        if department_codebook.exists():
            codebook_paths.append(department_codebook)
        profile_path = DEFAULT_SCHEDULE_PROFILE_DIR / f"{resolved_department}.json"
        schedule_profile = load_schedule_profile(profile_path if profile_path.exists() else None)
        applied_profile = str(profile_path) if profile_path.exists() else None
        return load_codebook(codebook_paths), [str(path) for path in codebook_paths if path.exists()], schedule_profile, applied_profile

    codebook_paths = codebook_paths_for_image(input_path, DEFAULT_CODEBOOK, DEFAULT_DEPARTMENT_CODEBOOK_DIR)
    profile_path = schedule_profile_path_for_image(input_path, DEFAULT_SCHEDULE_PROFILE_DIR)
    return (
        load_codebook(codebook_paths),
        [str(path) for path in codebook_paths if path.exists()],
        load_schedule_profile(profile_path),
        str(profile_path) if profile_path else None,
    )


def structured_records_to_schedule(
    records: list[dict[str, str]],
    source_path: Path,
    output_dir: Path,
    source_format: str,
    department_id: str | None = None,
) -> dict[str, Any]:
    from parse_schedule import (
        HARD_VALIDATION_COLUMNS,
        apply_profile_guided_repairs,
        apply_summary_guided_repairs,
        build_column_stats,
        render_html,
    )

    if not records:
        raise ValueError("No structured schedule rows found")

    start = time.time()
    output_dir.mkdir(parents=True, exist_ok=True)
    codebook, codebook_paths, schedule_profile, profile_path = load_codebook_and_profile(source_path, department_id)
    first = records[0]
    title = clean_value(first.get("title")) or source_path.stem
    active_days = parse_int(first.get("active_days")) or active_days_from_title(title)
    preview_asset = ensure_preview_asset(source_path, output_dir)
    employees: list[dict[str, Any]] = []

    for row_index, record in enumerate(records, start=1):
        role = clean_value(record.get("role"))
        name = clean_value(record.get("name"))
        display_name = clean_value(record.get("display_name")) or f"{role} {name}".strip()
        if not display_name:
            continue
        day_values = {str(day): clean_value(record.get(f"day_{day}")) for day in range(1, 32)}
        summary = {
            key: parse_int(record.get(SUMMARY_FIELD_NAMES[key]))
            for key in SUMMARY_KEYS
        }
        validation = validate_employee_summary(
            {day: value for day, value in day_values.items() if int(day) <= active_days},
            summary,
            codebook,
        )
        hard_mismatches = blocking_mismatches(validation, HARD_VALIDATION_COLUMNS)
        advisory_mismatches = {
            key: value for key, value in validation.mismatches.items() if key not in hard_mismatches
        }
        source_flag = f"structured-{source_format}"
        day_cells = {day: cell(value, source_flag) for day, value in day_values.items()}
        summary_cells = {
            key: cell(record.get(SUMMARY_FIELD_NAMES[key]), source_flag)
            for key in SUMMARY_KEYS
        }
        employee = {
            "group": clean_value(record.get("group")) or None,
            "group_display": clean_value(record.get("group")) or "\ubbf8\uc9c0\uc815",
            "section_index": row_index,
            "role": role,
            "name": name,
            "display_name": display_name,
            "previous": parse_int(record.get("previous")),
            "remaining": parse_int(record.get("remaining")),
            "name_cell": cell(display_name, source_flag),
            "previous_cell": cell(record.get("previous"), source_flag),
            "remaining_cell": cell(record.get("remaining"), source_flag),
            "days": day_values,
            "summary": summary,
            "validation": {
                "is_valid": not hard_mismatches and not validation.unknown_codes,
                "computed_summary": validation.computed,
                "printed_summary": validation.printed,
                "mismatches": validation.mismatches,
                "blocking_mismatches": hard_mismatches,
                "advisory_mismatches": advisory_mismatches,
                "unknown_codes": validation.unknown_codes,
            },
            "day_cells": day_cells,
            "summary_cells": summary_cells,
            "source_row_index": row_index,
            "warnings": [source_flag],
        }
        employees.append(employee)

    summary_guided_repairs = apply_summary_guided_repairs(employees, codebook, active_days)
    profile_guided_repairs = apply_profile_guided_repairs(employees, codebook, active_days, schedule_profile)
    invalid_rows = [employee for employee in employees if not employee["validation"]["is_valid"]]
    advisory_rows = [employee for employee in employees if employee["validation"].get("advisory_mismatches")]
    unknown_codes = sorted({code for employee in employees for code in employee["validation"]["unknown_codes"]})
    column_stats = build_column_stats(employees, codebook, active_days)
    column_validation = validate_column_profile(column_stats, schedule_profile)
    structured_quality_score = structured_quality(len(employees), invalid_rows, unknown_codes, column_validation)
    raw_rows = build_structured_raw_grid(employees, source_format)
    generated_at = dt.datetime.now().isoformat(timespec="seconds")
    schedule = {
        "document_type": "nurse_work_schedule",
        "source_format": source_format,
        "source_image": str(source_path),
        "image_asset": preview_asset,
        "generated_at": generated_at,
        "elapsed_seconds": round(time.time() - start, 2),
        "title": title,
        "printed_at": "",
        "active_days": active_days,
        "grid": {
            "rows": len(raw_rows),
            "columns": 40,
            "x_lines": [],
            "y_lines": [],
            "detection_params": {"source": source_format},
        },
        "quality": {
            "employee_count": len(employees),
            "valid_employee_rows": len(employees) - len(invalid_rows),
            "invalid_employee_rows": len(invalid_rows),
            "advisory_employee_rows": len(advisory_rows),
            "unknown_codes": unknown_codes,
            "codebook_paths": codebook_paths,
            "schedule_profile_path": profile_path,
            "column_profile_issue_count": column_validation["issue_count"],
            "column_profile_score": column_validation["profile_score"],
            "summary_guided_repairs": summary_guided_repairs,
            "profile_guided_repairs": profile_guided_repairs,
            "n_best_candidate_cells": 0,
            "n_best_candidates_total": 0,
            "structured_quality_score": structured_quality_score,
            "validation_rule": "Structured files bypass OCR and validate rows against D/E/N/slash/required/T totals.",
            "hard_validation_columns": HARD_VALIDATION_COLUMNS,
            "advisory_validation_columns": ["T"],
        },
        "columns": build_columns(),
        "column_stats": column_stats,
        "column_validation": column_validation,
        "employees": employees,
    }
    raw_grid = {
        "source_image": str(source_path),
        "image_asset": preview_asset,
        "generated_at": generated_at,
        "grid": schedule["grid"],
        "header_ocr": [],
        "cell_corrections": {},
        "rows": raw_rows,
    }
    schedule_path = output_dir / "schedule.json"
    raw_grid_path = output_dir / "schedule_raw_grid.json"
    html_path = output_dir / "index.html"
    schedule_path.write_text(json.dumps(schedule, ensure_ascii=False, indent=2), encoding="utf-8")
    raw_grid_path.write_text(json.dumps(raw_grid, ensure_ascii=False, indent=2), encoding="utf-8")
    render_html(schedule, raw_grid, html_path)
    return {
        "schedule": schedule_path,
        "raw_grid": raw_grid_path,
        "html": html_path,
        "employee_count": len(employees),
        "grid_rows": len(raw_rows),
        "grid_columns": 40,
        "elapsed_seconds": round(time.time() - start, 2),
    }


def structured_quality(
    employee_count: int,
    invalid_rows: Sequence[dict[str, Any]],
    unknown_codes: Sequence[str],
    column_validation: dict[str, Any],
) -> float:
    checks = max(employee_count + int(column_validation.get("checked_days", 0) or 0), 1)
    defects = len(invalid_rows) + len(unknown_codes) + int(column_validation.get("issue_count", 0) or 0)
    return round(max(0.0, 1.0 - defects / checks), 4)


def build_columns() -> list[dict[str, Any]]:
    return [
        {"index": 0, "key": "name", "label": "\uc131\uba85"},
        {"index": 1, "key": "previous", "label": "\uc804\ub2ec"},
        {"index": 2, "key": "remaining", "label": "\ub204\uc801"},
        *[{"index": day + 2, "key": str(day), "label": str(day)} for day in range(1, 32)],
        *[{"index": idx, "key": key, "label": key} for idx, key in zip(range(34, 40), SUMMARY_KEYS)],
    ]


def build_structured_raw_grid(employees: list[dict[str, Any]], source_format: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for employee in employees:
        values = [
            employee["display_name"],
            employee.get("previous"),
            employee.get("remaining"),
            *[employee["days"].get(str(day), "") for day in range(1, 32)],
            *[employee["summary"].get(key, "") for key in SUMMARY_KEYS],
        ]
        rows.append(
            {
                "row_index": employee["source_row_index"],
                "cells": [
                    {
                        "row_index": employee["source_row_index"],
                        "col_index": col_index,
                        "text": clean_value(value),
                        "raw_text": clean_value(value),
                        "confidence": 1.0,
                        "bbox": [],
                        "flags": [f"structured-{source_format}"],
                        "style": {},
                        "candidates": [],
                    }
                    for col_index, value in enumerate(values)
                ],
            }
        )
    return rows


def ensure_preview_asset(source_path: Path, output_dir: Path) -> str:
    if source_path.suffix.lower() in IMAGE_EXTENSIONS and source_path.exists():
        target = output_dir / source_path.name
        shutil.copy2(source_path, target)
        return target.name
    target = output_dir / "structured_source.png"
    if not target.exists():
        image = Image.new("RGB", (900, 180), "white")
        image.save(target)
    return target.name


def parse_pdf(path: Path, output_dir: Path, department_id: str | None = None) -> dict[str, Any]:
    digital_records = extract_pdf_table_records(path)
    if digital_records:
        return structured_records_to_schedule(digital_records, path, output_dir, "pdf-table", department_id)

    try:
        import pymupdf
    except ImportError as exc:
        raise RuntimeError("PyMuPDF is required to render scanned PDFs") from exc

    page_dir = output_dir / "_pdf_pages"
    page_dir.mkdir(parents=True, exist_ok=True)
    document = pymupdf.open(path)
    try:
        if len(document) == 0:
            raise ValueError("PDF has no pages")
        page = document[0]
        rendered = extract_first_embedded_pdf_image(document, page, path, page_dir)
        if rendered is None:
            # PIL image PDFs made from phone uploads preserve the source pixel size near 220dpi.
            # Rendering much higher than that resamples grid lines and hurts the cell OCR model.
            pixmap = page.get_pixmap(dpi=220)
            rendered = page_dir / f"{path.stem}.png"
            pixmap.save(rendered)
    finally:
        document.close()
    from parse_schedule import parse_schedule

    result = parse_schedule(
        rendered,
        output_dir,
        codebook_paths_for_image(Path(path.stem + ".jpeg"), DEFAULT_CODEBOOK, DEFAULT_DEPARTMENT_CODEBOOK_DIR),
        schedule_profile_path_for_image(Path(path.stem + ".jpeg"), DEFAULT_SCHEDULE_PROFILE_DIR),
    )
    schedule_path = Path(result["schedule"])
    schedule = json.loads(schedule_path.read_text(encoding="utf-8"))
    schedule["source_format"] = "pdf-scan"
    schedule["source_pdf"] = str(path)
    schedule_path.write_text(json.dumps(schedule, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def extract_first_embedded_pdf_image(document: Any, page: Any, path: Path, output_dir: Path) -> Path | None:
    images = page.get_images(full=True)
    if not images:
        return None
    best = max(images, key=lambda item: int(item[2]) * int(item[3]))
    xref = int(best[0])
    info = document.extract_image(xref)
    width = int(info.get("width") or 0)
    height = int(info.get("height") or 0)
    image_bytes = info.get("image")
    extension = str(info.get("ext") or "png").lower()
    if not image_bytes or width < 800 or height < 500:
        return None
    output = output_dir / f"{path.stem}.{extension}"
    output.write_bytes(image_bytes)
    return output


def extract_pdf_table_records(path: Path) -> list[dict[str, str]]:
    try:
        import pdfplumber
    except ImportError:
        return []
    records: list[dict[str, str]] = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables() or []:
                if len(table) < 2:
                    continue
                headers = [clean_value(value) for value in table[0]]
                if not all(header in headers for header in ["title", "display_name", "day_1"]):
                    continue
                for row in table[1:]:
                    records.append(
                        {
                            headers[index]: clean_value(value)
                            for index, value in enumerate(row)
                            if index < len(headers)
                        }
                    )
    return records


def parse_any_upload(path: Path, output_dir: Path, department_id: str | None = None) -> dict[str, Any]:
    kind = detect_file_kind(path)
    if kind == "image":
        from parse_schedule import parse_schedule

        return parse_schedule(
            path,
            output_dir,
            codebook_paths_for_image(path, DEFAULT_CODEBOOK, DEFAULT_DEPARTMENT_CODEBOOK_DIR),
            schedule_profile_path_for_image(path, DEFAULT_SCHEDULE_PROFILE_DIR),
        )
    if kind == "csv":
        return structured_records_to_schedule(read_csv_records(path), path, output_dir, "csv", department_id)
    if kind == "xls-html":
        return structured_records_to_schedule(read_xls_html_records(path), path, output_dir, "xls-html", department_id)
    if kind == "xlsx":
        return structured_records_to_schedule(read_xlsx_records(path), path, output_dir, "xlsx", department_id)
    if kind == "pdf":
        return parse_pdf(path, output_dir, department_id)
    if kind == "xls-binary":
        raise ValueError("Legacy binary .xls is not supported yet. Please upload CSV or XLSX.")
    raise ValueError(f"Unsupported upload file type: {path.suffix or 'unknown'}")


def parse_structured_file(path: Path) -> list[dict[str, str]]:
    kind = detect_file_kind(path)
    readers = {
        "csv": read_csv_records,
        "xls-html": read_xls_html_records,
        "xlsx": read_xlsx_records,
    }
    if kind not in readers:
        raise ValueError(f"Not a structured table file: {kind}")
    return readers[kind](path)


def normalize_records(records: Iterable[dict[str, str]]) -> list[dict[str, str]]:
    return [
        {header: clean_value(record.get(header, "")) for header in STRUCTURED_HEADERS}
        for record in records
    ]
