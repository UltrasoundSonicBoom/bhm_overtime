from __future__ import annotations

import csv
import html
import json
from pathlib import Path
from typing import Any

from PIL import Image


REQUIRED = "\ud544"
SUMMARY_KEYS = ["D", "E", "N", "/", REQUIRED, "T"]
SUMMARY_FIELD_NAMES = {
    "D": "summary_D",
    "E": "summary_E",
    "N": "summary_N",
    "/": "summary_slash",
    REQUIRED: "summary_required",
    "T": "summary_T",
}
DAY_FIELD_NAMES = [f"day_{day}" for day in range(1, 32)]
STRUCTURED_HEADERS = [
    "document_id",
    "source_image",
    "title",
    "active_days",
    "department_id",
    "group",
    "role",
    "name",
    "display_name",
    "previous",
    "remaining",
    *DAY_FIELD_NAMES,
    *[SUMMARY_FIELD_NAMES[key] for key in SUMMARY_KEYS],
]


def document_id_from_schedule(schedule: dict[str, Any], fallback: str = "schedule") -> str:
    source = str(schedule.get("source_image") or schedule.get("image_asset") or fallback)
    return Path(source).stem


def department_id_from_document_id(document_id: str) -> str | None:
    digits = []
    for char in document_id:
        if char.isdigit():
            digits.append(char)
            continue
        break
    return "".join(digits) or None


def value_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def schedule_to_records(schedule: dict[str, Any]) -> list[dict[str, str]]:
    document_id = document_id_from_schedule(schedule)
    department_id = department_id_from_document_id(document_id) or ""
    records: list[dict[str, str]] = []
    for employee in schedule.get("employees", []):
        record: dict[str, str] = {
            "document_id": document_id,
            "source_image": value_text(schedule.get("source_image")),
            "title": value_text(schedule.get("title") or document_id),
            "active_days": value_text(schedule.get("active_days") or ""),
            "department_id": department_id,
            "group": value_text(employee.get("group") or ""),
            "role": value_text(employee.get("role") or ""),
            "name": value_text(employee.get("name") or ""),
            "display_name": value_text(employee.get("display_name") or ""),
            "previous": value_text(employee.get("previous")),
            "remaining": value_text(employee.get("remaining")),
        }
        for day in range(1, 32):
            record[f"day_{day}"] = value_text(employee.get("days", {}).get(str(day), ""))
        for key in SUMMARY_KEYS:
            record[SUMMARY_FIELD_NAMES[key]] = value_text(employee.get("summary", {}).get(key, ""))
        records.append(record)
    return records


def export_csv(records: list[dict[str, str]], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=STRUCTURED_HEADERS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)
    return path


def export_xls_html(records: list[dict[str, str]], path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    header_html = "".join(f"<th>{html.escape(header)}</th>" for header in STRUCTURED_HEADERS)
    rows = []
    for record in records:
        cells = "".join(
            f"<td>{html.escape(value_text(record.get(header)))}</td>"
            for header in STRUCTURED_HEADERS
        )
        rows.append(f"<tr>{cells}</tr>")
    path.write_text(
        "<!doctype html><html><head><meta charset=\"utf-8\"></head><body>"
        f"<table><thead><tr>{header_html}</tr></thead><tbody>{''.join(rows)}</tbody></table>"
        "</body></html>",
        encoding="utf-8",
    )
    return path


def export_xlsx(records: list[dict[str, str]], path: Path) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to export xlsx files") from exc

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "schedule"
    sheet.append(STRUCTURED_HEADERS)
    for record in records:
        sheet.append([record.get(header, "") for header in STRUCTURED_HEADERS])
    sheet.freeze_panes = "L2"
    header_fill = PatternFill("solid", fgColor="D9E2EC")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    for column_cells in sheet.columns:
        max_length = max(len(value_text(cell.value)) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 8), 18)
    workbook.save(path)
    return path


def export_pdf_from_source_image(schedule: dict[str, Any], path: Path) -> Path:
    source_image = Path(str(schedule.get("source_image") or ""))
    if not source_image.exists():
        raise FileNotFoundError(f"source image not found for PDF export: {source_image}")
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        import pymupdf

        with Image.open(source_image) as image:
            width, height = image.size
        document = pymupdf.open()
        page = document.new_page(width=width, height=height)
        page.insert_image(pymupdf.Rect(0, 0, width, height), filename=str(source_image))
        document.save(path)
        document.close()
        return path
    except ImportError:
        pass

    image = Image.open(source_image).convert("RGB")
    image.save(path, "PDF", resolution=220.0)
    return path


def export_schedule_formats(
    schedule_path: Path,
    output_dir: Path,
    include_pdf: bool = True,
) -> dict[str, Path]:
    schedule = json.loads(Path(schedule_path).read_text(encoding="utf-8"))
    document_id = document_id_from_schedule(schedule, Path(schedule_path).parent.name)
    output_dir.mkdir(parents=True, exist_ok=True)
    records = schedule_to_records(schedule)
    exports = {
        "csv": export_csv(records, output_dir / f"{document_id}.csv"),
        "xls": export_xls_html(records, output_dir / f"{document_id}.xls"),
        "xlsx": export_xlsx(records, output_dir / f"{document_id}.xlsx"),
    }
    if include_pdf:
        exports["pdf"] = export_pdf_from_source_image(schedule, output_dir / f"{document_id}.pdf")
    return exports


def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="Export a parsed schedule to CSV, Excel, and scanned PDF fixtures.")
    parser.add_argument("schedule", type=Path)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--no-pdf", action="store_true")
    args = parser.parse_args()
    exports = export_schedule_formats(args.schedule, args.output_dir, include_pdf=not args.no_pdf)
    print(json.dumps({key: str(value) for key, value in exports.items()}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
